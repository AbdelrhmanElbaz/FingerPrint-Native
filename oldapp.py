import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import xlrd
import re
import io
import json
import math
import hashlib
import os
import shutil
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# session_manager — إدارة الجلسات والملفات المحفوظة
# ============================================================

def _get_data_root() -> Path:
    import sys
    if getattr(sys, "frozen", False):
        base = Path(sys.executable).parent
    else:
        base = Path(__file__).parent
    root = base / "data"
    root.mkdir(exist_ok=True)
    return root

def _get_company_dir(company: str) -> Path:
    d = _get_data_root() / company
    d.mkdir(exist_ok=True)
    return d

def _get_month_dir(company: str, year: int, month: int) -> Path:
    d = _get_company_dir(company) / str(year) / f"{month:02d}"
    d.mkdir(parents=True, exist_ok=True)
    return d

def _session_file(company: str) -> Path:
    return _get_company_dir(company) / "last_session.json"

def _defaults_file(company: str) -> Path:
    return _get_company_dir(company) / "employees_defaults.json"

def list_companies() -> list:
    root = _get_data_root()
    return sorted([d.name for d in root.iterdir() if d.is_dir()])

def save_imported_file(file_bytes: bytes, company: str, year: int, month: int) -> Path:
    d = _get_month_dir(company, year, month)
    base_name = f"{company}_{year}_{month:02d}"
    work_path = d / f"{base_name}.xlsx"
    orig_path = d / f"{base_name}_original.bin"
    if not orig_path.exists():
        orig_path.write_bytes(file_bytes)
    work_path.write_bytes(file_bytes)
    session = {
        "company": company,
        "year": year,
        "month": month,
        "work_file": str(work_path),
        "orig_file": str(orig_path),
    }
    _session_file(company).write_text(json.dumps(session, ensure_ascii=False), encoding="utf-8")
    (_get_data_root() / "last_company.txt").write_text(company, encoding="utf-8")
    return work_path

def load_last_session():
    last_co_file = _get_data_root() / "last_company.txt"
    if not last_co_file.exists():
        return None
    company = last_co_file.read_text(encoding="utf-8").strip()
    sf = _session_file(company)
    if not sf.exists():
        return None
    try:
        return json.loads(sf.read_text(encoding="utf-8"))
    except Exception:
        return None

def save_work_file(file_bytes: bytes, company: str, year: int, month: int):
    d = _get_month_dir(company, year, month)
    work_path = d / f"{company}_{year}_{month:02d}.xlsx"
    work_path.write_bytes(file_bytes)

def load_work_file_bytes(company: str, year: int, month: int):
    d = _get_month_dir(company, year, month)
    work_path = d / f"{company}_{year}_{month:02d}.xlsx"
    if work_path.exists():
        return work_path.read_bytes()
    return None

def load_original_file_bytes(company: str, year: int, month: int):
    d = _get_month_dir(company, year, month)
    orig_path = d / f"{company}_{year}_{month:02d}_original.bin"
    if orig_path.exists():
        return orig_path.read_bytes()
    return None

def save_employee_defaults(company: str, rates: dict):
    existing = load_employee_defaults(company)
    existing.update({k: v for k, v in rates.items() if v and v > 0})
    _defaults_file(company).write_text(
        json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8"
    )

def load_employee_defaults(company: str) -> dict:
    f = _defaults_file(company)
    if not f.exists():
        return {}
    try:
        return json.loads(f.read_text(encoding="utf-8"))
    except Exception:
        return {}

# ── إعدادات كل ملف (شهر) منفصلة ──────────────────────────────
_ARABIC_MONTHS = [
    '', 'يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو',
    'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر'
]

def _month_settings_file(company: str, year: int, month: int) -> Path:
    return _get_month_dir(company, year, month) / "session_settings.json"

def _get_file_state_file(company: str, year: int, month: int) -> Path:
    """مسار ملف الحالة الشاملة لملف شهر معيّن"""
    return _get_month_dir(company, year, month) / "file_state.json"

def save_month_settings(company: str, year: int, month: int, settings: dict):
    """يحفظ إعدادات الشيفتات (cutoff/saturate) الخاصة بملف شهر معيّن."""
    try:
        _month_settings_file(company, year, month).write_text(
            json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    except Exception:
        pass

def load_month_settings(company: str, year: int, month: int) -> dict:
    """يرجّع إعدادات ملف شهر معيّن، أو {} لو الملف ده بيُفتح لأول مرة."""
    f = _month_settings_file(company, year, month)
    if not f.exists():
        return {}
    try:
        return json.loads(f.read_text(encoding="utf-8"))
    except Exception:
        return {}

def save_file_state(company: str, year: int, month: int, state: dict):
    """
    يحفظ حالة الملف الشاملة:
    - corrections: التصحيحات على البصمات
    - day_overrides: تعديلات الأيام
    - fingerprint_types: نوع البصمة لكل موظف (punch/not_punch)
    - hourly_rates: أسعار الساعات
    - missing_hours: الساعات الناقصة المضافة
    """
    try:
        _get_file_state_file(company, year, month).write_text(
            json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    except Exception as e:
        print(f"❌ خطأ عند حفظ حالة الملف: {e}")

def load_file_state(company: str, year: int, month: int) -> dict:
    """
    يحمّل حالة الملف الشاملة (أو {} إذا لم يكن الملف موجوداً)
    """
    f = _get_file_state_file(company, year, month)
    if not f.exists():
        return {}
    try:
        return json.loads(f.read_text(encoding="utf-8"))
    except Exception as e:
        print(f"❌ خطأ عند تحميل حالة الملف: {e}")
        return {}

# ── إدارة الشركات والملفات (تبويب "📁 ملفاتي") ─────────────────
def list_company_tree() -> dict:
    """
    يرجّع شجرة كل الملفات المحفوظة:
    { "MYGYM": { 2026: [6, 5] }, "OTHERGYM": { 2026: [6] } }
    الشهور مرتّبة تنازليًا (الأحدث أولاً).
    """
    tree = {}
    for company in list_companies():
        co_dir = _get_company_dir(company)
        years = {}
        for year_dir in co_dir.iterdir():
            if not year_dir.is_dir() or not year_dir.name.isdigit():
                continue
            months = []
            for month_dir in year_dir.iterdir():
                if not month_dir.is_dir() or not month_dir.name.isdigit():
                    continue
                base = f"{company}_{year_dir.name}_{month_dir.name}"
                if (month_dir / f"{base}.xlsx").exists():
                    months.append(int(month_dir.name))
            if months:
                years[int(year_dir.name)] = sorted(months, reverse=True)
        if years:
            tree[company] = years
    return tree

def delete_month_file(company: str, year: int, month: int):
    """يحذف ملف شهر معيّن (الشغل + الأصلي + الإعدادات) وينظّف الفولدرات الفاضية."""
    d = _get_month_dir(company, year, month)
    if d.exists():
        shutil.rmtree(d, ignore_errors=True)
    year_dir = _get_company_dir(company) / str(year)
    try:
        if year_dir.exists() and not any(year_dir.iterdir()):
            year_dir.rmdir()
    except Exception:
        pass
    # لو كان هو الملف المفتوح حاليًا في last_session، نمسح المرجع
    sf = _session_file(company)
    if sf.exists():
        try:
            sess = json.loads(sf.read_text(encoding="utf-8"))
            if sess.get("year") == year and sess.get("month") == month:
                sf.unlink(missing_ok=True)
        except Exception:
            pass

def delete_company(company: str):
    """يحذف شركة كاملة بكل ملفاتها."""
    co_dir = _get_company_dir(company)
    if co_dir.exists():
        shutil.rmtree(co_dir, ignore_errors=True)
    last_co_file = _get_data_root() / "last_company.txt"
    if last_co_file.exists() and last_co_file.read_text(encoding="utf-8").strip() == company:
        last_co_file.unlink(missing_ok=True)

def auto_save_file_state():
    """
    دالة مساعدة لحفظ حالة الملف تلقائياً من session_state
    تُستدعى بعد أي تعديل مهم
    """
    if not st.session_state.get('_current_company') or not st.session_state.get('_current_year') or not st.session_state.get('_current_month'):
        return
    
    company = st.session_state['_current_company']
    year = st.session_state['_current_year']
    month = st.session_state['_current_month']
    
    state = {
        'corrections': st.session_state.get('corrections', {}),
        'day_overrides': st.session_state.get('day_overrides', {}),
        'fingerprint_types': st.session_state.get('fingerprint_types', {}),
        'hourly_rates': st.session_state.get('saved_rates', {}),
        'missing_hours': st.session_state.get('missing_hours', {}),
    }
    
    save_file_state(company, year, month, state)

def rename_company(old_name: str, new_name: str) -> bool:
    """يغيّر اسم شركة (فولدرها). يرجّع False لو الاسم الجديد مستخدم بالفعل."""
    old_dir = _get_data_root() / old_name
    new_dir = _get_data_root() / new_name
    if new_dir.exists() or not old_dir.exists():
        return False
    old_dir.rename(new_dir)
    last_co_file = _get_data_root() / "last_company.txt"
    if last_co_file.exists() and last_co_file.read_text(encoding="utf-8").strip() == old_name:
        last_co_file.write_text(new_name, encoding="utf-8")
    return True

# ============================================================
# نهاية session_manager
# ============================================================

st.set_page_config(page_title="نظام الحضور والرواتب", layout="wide", page_icon="💼")

# ── توافق مع الإصدارات المختلفة من Streamlit لخاصية الـ Fragment ──
# الـ fragment بيخلي أي تفاعل (زي طي/فتح عنصر في شجرة الملفات) يعيد تشغيل
# الجزء ده بس من الكود، مش الصفحة كلها (اللي بتتضمن تحليل ملف الإكسل وحساب الرواتب)
# فالتبديل بين الملفات في الشجرة بيبقى سريع فورًا بدل ما يعلّق الصفحة.
if hasattr(st, "fragment"):
    _st_fragment = st.fragment
elif hasattr(st, "experimental_fragment"):
    _st_fragment = st.experimental_fragment
else:
    def _st_fragment(func=None, **_kwargs):
        if func is None:
            def _deco(f):
                return f
            return _deco
        return func

def _full_rerun():
    """يعمل rerun لكامل الصفحة (مش الـ fragment بس) — يُستخدم لأي حدث لازم
    يغيّر بيانات الملف المفتوح فعليًا (فتح/حذف/تسمية)، على عكس مجرد طي/فتح شجرة."""
    try:
        st.rerun(scope="app")
    except TypeError:
        st.rerun()

# رقم إصدار هذا البرنامج - عدّله يدويًا مع كل Build جديد ترسله للعملاء
APP_VERSION = "1.0.0"

# ============================================================
# نسخة أوفلاين بالكامل - بدون تسجيل دخول أو ترخيص أو اتصال بالإنترنت
# ============================================================
if "license_client" not in st.session_state:
    st.session_state["license_client"] = {
        "client_id": "offline-local",
        "client_name": "User001",
        "is_active": True,
    }
# نهاية النسخة الأوفلاين - بقية التطبيق الأصلي يبدأ من هنا

# ── زر Sidebar ثابت — يُحقن مرة واحدة في document الأب عبر iframe ──
# ملاحظة: محاط بـ try/catch لدعم المتصفحات التي تقيّد وصول الـ iframe للـ parent
# (Permissions-Policy / Feature-Policy) — في حال الفشل يُعاد تفعيل الزر الأصلي تلقائياً
components.html("""
<script>
(function() {
    // ── الوصول للـ parent محاط بـ try/catch
    // لأن بعض المتصفحات أو إعدادات Permissions-Policy قد تمنعه
    var p, doc;
    try {
        p   = window.parent;
        doc = p.document;
        // اختبار سريع: لو المتصفح منع الوصول سيرمي خطأ هنا
        void doc.readyState;
    } catch(secErr) {
        // ── Fallback: المتصفح منع الوصول للـ parent ──
        // نُعيد إظهار الزر الأصلي لستريملت حتى لا تختفي وظيفة فتح/إغلاق السايدبار
        console.warn('[SidebarToggle] لا يمكن الوصول للـ parent frame:', secErr);
        try {
            var styleEl = document.createElement('style');
            styleEl.textContent =
                '[data-testid="stSidebarCollapseButton"],' +
                '[data-testid="collapsedControl"] {' +
                '  opacity: 1 !important;' +
                '  pointer-events: auto !important;' +
                '  position: static !important;' +
                '  top: auto !important;' +
                '}';
            document.head.appendChild(styleEl);
        } catch(e) { /* لا يوجد ما يمكن فعله */ }
        return; // نوقف تنفيذ باقي الكود
    }

    function doInject() {
        var sidebar = doc.querySelector('[data-testid="stSidebar"]');
        var existingBtn = doc.getElementById('_sb_toggle_btn');

        // ── مفيش Sidebar حاليًا (مثلاً: شاشة تسجيل الدخول) ──
        // نحذف أي زر قديم متبقي عشان ميفضلش عائمًا بلا فائدة
        if (!sidebar) {
            if (existingBtn) existingBtn.remove();
            return;
        }

        // ── فيه Sidebar، لكن الزر القديم (لو موجود) متربوط بعنصر مختلف ──
        // (يحصل بعد تسجيل خروج ثم تسجيل دخول من جديد - الـ Sidebar بيتبني من الصفر)
        if (existingBtn) {
            if (existingBtn._boundSidebar === sidebar) return; // كل شيء سليم بالفعل
            existingBtn.remove(); // عنصر قديم متعلّق بـ Sidebar غير موجود في DOM - احذفه وأعد البناء
        }

        /* ── الزر ── */
        var btn = doc.createElement('button');
        btn.id = '_sb_toggle_btn';
        btn._boundSidebar = sidebar;
        Object.assign(btn.style, {
            position:     'fixed',
            top:          '50%',
            right:        '0px',
            transform:    'translateY(-50%)',
            zIndex:       '2147483647',
            width:        '28px',
            height:       '56px',
            background:   '#181715',
            color:        '#faf9f5',
            border:       '1px solid #cc785c',
            borderRight:  'none',
            borderRadius: '8px 0 0 8px',
            cursor:       'pointer',
            fontSize:     '20px',
            fontWeight:   '700',
            display:      'flex',
            alignItems:   'center',
            justifyContent: 'center',
            transition:   'background .2s, border-color .2s, right .28s cubic-bezier(.4,0,.2,1)',
            padding:      '0',
            lineHeight:   '1',
        });

        function isOpen() {
            return sidebar.getAttribute('aria-expanded') !== 'false';
        }
        function updateBtn() {
            var open = isOpen();
            btn.textContent = open ? '‹' : '›';
            btn.title       = open ? 'إغلاق القائمة' : 'فتح القائمة';
            btn.style.right = open ? 'calc(21rem)' : '0px';
        }
        updateBtn();

        btn.addEventListener('mouseenter', function() {
            btn.style.background   = '#2a3050';
            btn.style.borderColor  = '#7a80b0';
        });
        btn.addEventListener('mouseleave', function() {
            btn.style.background   = '#1e2438';
            btn.style.borderColor  = '#4a5070';
        });

        btn.addEventListener('click', function() {
            try {
                /* اضغط على الزر الأصلي المخفي */
                var native =
                    doc.querySelector('[data-testid="stSidebarCollapseButton"] button') ||
                    doc.querySelector('[data-testid="collapsedControl"] button');
                if (native) {
                    native.click();
                    p.setTimeout(updateBtn, 60);
                    p.setTimeout(updateBtn, 350);
                }
            } catch(clickErr) {
                console.warn('[SidebarToggle] فشل النقر على الزر الأصلي:', clickErr);
            }
        });

        try {
            /* راقب تغيير الحالة */
            new p.MutationObserver(updateBtn)
                .observe(sidebar, { attributes: true, attributeFilter: ['aria-expanded'] });

            doc.body.appendChild(btn);
        } catch(injectErr) {
            console.warn('[SidebarToggle] فشل حقن الزر:', injectErr);
        }
    }

    // راقب الصفحة باستمرار (وليس مرة واحدة فقط) عشان يتكيّف تلقائيًا
    // مع أي تسجيل خروج/دخول لاحق بدون الحاجة لعمل Refresh يدوي
    try {
        if (doc.readyState === 'loading') {
            doc.addEventListener('DOMContentLoaded', function() { p.setTimeout(doInject, 200); });
        } else {
            p.setTimeout(doInject, 200);
        }
        if (!p._sbToggleWatcherStarted) {
            p._sbToggleWatcherStarted = true;
            p.setInterval(doInject, 700);
        }
    } catch(domErr) {
        console.warn('[SidebarToggle] فشل الانتظار لـ DOM:', domErr);
    }
})();
</script>
""", height=0)

st.markdown("""
<style>
/* ── Claude-lite editorial font stack: Copernicus-style serif + StyreneB/Inter sans ── */
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans+Arabic:wght@300;400;500;600;700&display=swap');
@import url('https://fonts.googleapis.com/css2?family=Cormorant+Garamond:wght@500;600&family=Inter:wght@400;500;600;700&display=swap');
@import url('https://fonts.googleapis.com/icon?family=Material+Icons');

/* ════════════════════════════════════════════════════════════════
   DESIGN SYSTEM — Claude-lite Editorial (cream / coral / dark navy)
   Font: IBM Plex Sans Arabic (Arabic body/UI) + Cormorant Garamond (display)
   ════════════════════════════════════════════════════════════════ */

/* ── 1. Global Canvas ── */
html, body,
.stApp,
[data-testid="stAppViewContainer"],
[data-testid="stMain"],
.main,
.block-container {
    background-color: #faf9f5 !important;
    color: #141413 !important;
    font-family: 'IBM Plex Sans Arabic', 'StyreneB', 'Inter', system-ui, "Segoe UI", Roboto, Helvetica, Arial, sans-serif !important;
}

.block-container {
    padding-top: 2rem !important;
    padding-bottom: 5rem !important;
    max-width: 1280px !important;
}

/* ── 2. RTL Direction ── */
html, body, .stApp, [data-testid="stAppViewContainer"],
[data-testid="stHeader"], [data-testid="stSidebar"],
[data-testid="stMain"], .main, .block-container {
    direction: rtl;
}

h1, h2, h3, h4, h5, h6, p, span, label, li,
.stMarkdown, .stMarkdown p, .stMarkdown li,
.stCaption, [data-testid="stCaptionContainer"],
[data-testid="stMetricLabel"], [data-testid="stMetricValue"],
[data-testid="stMetricDelta"], .stAlert, .stAlert p,
.stException, .stTooltipIcon {
    direction: rtl;
    text-align: right;
}

.stTextInput input, .stNumberInput input, .stTextArea textarea,
.stSelectbox div[data-baseweb="select"] *,
.stMultiSelect div[data-baseweb="select"] * {
    direction: rtl;
    text-align: right;
}

.stTextInput label, .stNumberInput label, .stTextArea label,
.stSelectbox label, .stMultiSelect label,
.stCheckbox, .stCheckbox label,
.stRadio, .stRadio label,
.stFileUploader label, .stFileUploader, .stDateInput label {
    direction: rtl;
    text-align: right;
    width: 100%;
}

.stButton button, .stDownloadButton button, .stFormSubmitButton button {
    direction: rtl;
}

[data-baseweb="popover"] li, [data-baseweb="menu"] li {
    direction: rtl;
    text-align: right;
}

[data-baseweb="tab-list"] { direction: rtl; }
[data-baseweb="tab"] { direction: rtl; }

.streamlit-expanderHeader, [data-testid="stExpander"] summary {
    direction: rtl;
    text-align: right;
}

[data-testid="stFileUploaderDropzone"] { direction: rtl; }
[data-testid="stFileUploaderDropzoneInstructions"] { text-align: right; }
[data-testid="stHorizontalBlock"] { direction: rtl; }
[data-testid="stDataFrame"], [data-testid="stDataFrameResizable"] { direction: rtl; }

/* ── 3. Header Bar ── */
[data-testid="stHeader"] {
    background-color: #faf9f5 !important;
    border-bottom: 1px solid #e6dfd8 !important;
    z-index: 99 !important;
}

/* ════════════════════════════════════════════════════════════
   4. SIDEBAR — RTL right-side drawer
   ════════════════════════════════════════════════════════════ */

/* ── إخفاء الزرارين الأصليين — زرنا المُحقن بالـ JS يتولى الأمر ── */
[data-testid="stSidebarCollapseButton"],
[data-testid="collapsedControl"] {
    opacity: 0 !important;
    pointer-events: none !important;
    position: fixed !important;
    top: -9999px !important;
}


/* ── Drawer panel ── */
[data-testid="stSidebar"] {
    position: fixed !important;
    top: 0 !important;
    right: 0 !important;
    left: auto !important;
    height: 100vh !important;
    width: 21rem !important;
    background-color: #faf9f5 !important;
    border-left: 1px solid #e6dfd8 !important;
    border-right: none !important;
    z-index: 200 !important;
    overflow-y: auto !important;
    overflow-x: hidden !important;
    transform: translateX(0) !important;
    transition: transform 0.28s cubic-bezier(.4,0,.2,1),
                visibility 0.28s !important;
    visibility: visible !important;
}

[data-testid="stSidebar"][aria-expanded="false"] {
    transform: translateX(100%) !important;
    visibility: hidden !important;
    pointer-events: none !important;
    border: none !important;
}

[data-testid="stSidebarHeader"] {
    overflow: hidden !important;
}

/* ── إخفاء resize handle ── */
div[style*="cursor: col-resize"] {
    display: none !important;
}

/* ── Material Icons ── */
[data-testid="stIconMaterial"] {
    font-size: 0 !important;
    color: transparent !important;
    line-height: 0 !important;
    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;
    width: 20px !important;
    height: 20px !important;
}

[data-testid="stSidebar"] .stMarkdown h2,
[data-testid="stSidebar"] .stMarkdown h3 {
    color: #cc785c !important;
    font-size: 0.78rem !important;
    font-weight: 600 !important;
    letter-spacing: 0.08em !important;
    text-transform: uppercase !important;
    margin-bottom: 0.6rem !important;
}

[data-testid="stSidebar"] label,
[data-testid="stSidebar"] .stCaption,
[data-testid="stSidebar"] p {
    color: #000 !important;
    font-size: 0.82rem !important;
}

[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h4,
[data-testid="stSidebar"] h5,
[data-testid="stSidebar"] h6,
[data-testid="stSidebar"] span,
[data-testid="stSidebar"] li,
[data-testid="stSidebar"] .stMarkdown {
    color: #000 !important;
}

/* ── 5. Page Title ── */
h1 {
    font-family: 'Cormorant Garamond', 'IBM Plex Sans Arabic', serif !important;
    font-size: 1.9rem !important;
    font-weight: 600 !important;
    color: #141413 !important;
    letter-spacing: -0.02em !important;
    margin-bottom: 0.25rem !important;
}

/* ── 6. Section Headers ── */
h2 {
    font-family: 'Cormorant Garamond', 'IBM Plex Sans Arabic', serif !important;
    font-size: 1.3rem !important;
    font-weight: 600 !important;
    color: #252523 !important;
    border-bottom: 1px solid #e6dfd8 !important;
    padding-bottom: 0.5rem !important;
    margin-top: 2rem !important;
    margin-bottom: 1rem !important;
    letter-spacing: -0.01em !important;
}

h3 {
    font-size: 0.9rem !important;
    font-weight: 600 !important;
    color: #6c6a64 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.06em !important;
    margin-top: 1.2rem !important;
    margin-bottom: 0.6rem !important;
}

/* ── 7. Metrics / KPI Cards ── */
[data-testid="stMetric"] {
    background: #efe9de !important;
    border: 1px solid #e6dfd8 !important;
    border-radius: 12px !important;
    padding: 1rem 1.25rem !important;
    transition: border-color 0.2s ease !important;
}

[data-testid="stMetric"]:hover {
    border-color: #cc785c !important;
}

[data-testid="stMetricLabel"] {
    color: #6c6a64 !important;
    font-size: 0.75rem !important;
    font-weight: 500 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.07em !important;
}

[data-testid="stMetricValue"] {
    color: #141413 !important;
    font-size: 1.9rem !important;
    font-weight: 700 !important;
    letter-spacing: -0.03em !important;
    line-height: 1.1 !important;
}

[data-testid="stMetricDelta"] {
    font-size: 0.78rem !important;
}

/* ── 8. Buttons ── */
.stButton > button {
    background: #faf9f5 !important;
    color: #141413 !important;
    border: 1px solid #e6dfd8 !important;
    border-radius: 8px !important;
    font-size: 0.84rem !important;
    font-weight: 500 !important;
    padding: 0.45rem 1rem !important;
    transition: all 0.15s ease !important;
    box-shadow: none !important;
}

.stButton > button:hover {
    background: #efe9de !important;
    border-color: #cc785c !important;
    color: #141413 !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 12px #efe9de !important;
}

/* Primary button */
.stButton > button[kind="primary"],
[data-testid="baseButton-primary"] {
    background: #cc785c !important;
    color: #000 !important;
    border: 1px solid #cc785c !important;
    font-weight: 600 !important;
    box-shadow: 0 2px 8px rgba(204,120,92,0.25) !important;
}

.stButton > button[kind="primary"]:hover,
[data-testid="baseButton-primary"]:hover {
    background: #a9583e !important;
    border-color: #a9583e !important;
    box-shadow: 0 4px 16px rgba(204,120,92,0.4) !important;
    transform: translateY(-1px) !important;
}

/* Download button */
.stDownloadButton > button {
    background: #5db872 !important;
    color: #000 !important;
    border: 1px solid #5db872 !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    font-size: 0.84rem !important;
    box-shadow: 0 2px 8px rgba(93,184,114,0.25) !important;
    transition: all 0.15s ease !important;
}

.stDownloadButton > button:hover {
    background: #4ea562 !important;
    box-shadow: 0 4px 16px rgba(93,184,114,0.4) !important;
    transform: translateY(-1px) !important;
}

/* ── 9. Inputs ── */
.stTextInput > div > div > input,
.stNumberInput > div > div > input,
.stTextArea > div > textarea {
    background: #faf9f5 !important;
    border: 1px solid #e6dfd8 !important;
    border-radius: 8px !important;
    color: #141413 !important;
    font-size: 0.85rem !important;
    padding: 0.5rem 0.85rem !important;
    transition: border-color 0.15s ease, box-shadow 0.15s ease !important;
}

.stTextInput > div > div > input:focus,
.stNumberInput > div > div > input:focus,
.stTextArea > div > textarea:focus {
    border-color: #cc785c !important;
    box-shadow: 0 0 0 3px rgba(204,120,92,0.15) !important;
    outline: none !important;
}

/* Selectbox */
.stSelectbox > div > div {
    background: #faf9f5 !important;
    border: 1px solid #e6dfd8 !important;
    border-radius: 8px !important;
    color: #141413 !important;
}

.stSelectbox > div > div:focus-within {
    border-color: #cc785c !important;
    box-shadow: 0 0 0 3px rgba(204,120,92,0.15) !important;
}

/* ── 10. Labels & Captions ── */
label {
    color: #6c6a64 !important;
    font-size: 0.8rem !important;
    font-weight: 500 !important;
    margin-bottom: 0.25rem !important;
}

.stCaption, [data-testid="stCaptionContainer"] {
    color: #8e8b82 !important;
    font-size: 0.75rem !important;
}

/* ── 11. Alerts ── */
/* Info */
[data-testid="stInfo"] {
    background: rgba(93,184,166,0.10) !important;
    border: 1px solid rgba(93,184,166,0.3) !important;
    border-radius: 8px !important;
    color: #2f6f63 !important;
}

[data-testid="stInfo"] p { color: #2f6f63 !important; }

/* Success */
[data-testid="stSuccess"],
div[data-baseweb="notification"][kind="positive"] {
    background: rgba(93,184,114,0.10) !important;
    border: 1px solid rgba(93,184,114,0.3) !important;
    border-radius: 8px !important;
    color: #3c7a4c !important;
}

[data-testid="stSuccess"] p { color: #3c7a4c !important; }

/* Warning */
[data-testid="stWarning"],
div[data-baseweb="notification"][kind="warning"] {
    background: rgba(212,160,23,0.10) !important;
    border: 1px solid rgba(212,160,23,0.3) !important;
    border-radius: 8px !important;
    color: #8a6a10 !important;
}

[data-testid="stWarning"] p { color: #8a6a10 !important; }

/* Error */
[data-testid="stError"],
div[data-baseweb="notification"][kind="negative"] {
    background: rgba(198,69,69,0.08) !important;
    border: 1px solid rgba(198,69,69,0.25) !important;
    border-radius: 8px !important;
    color: #c64545 !important;
}

[data-testid="stError"] p { color: #c64545 !important; }

/* ── 12. Expanders ── */
[data-testid="stExpander"] {
    background: #efe9de !important;
    border: 1px solid #e6dfd8 !important;
    border-radius: 12px !important;
    margin-bottom: 0.6rem !important;
    overflow: hidden !important;
}

[data-testid="stExpander"]:hover {
    border-color: #cc785c !important;
}

.streamlit-expanderHeader,
[data-testid="stExpander"] summary {
    background: #efe9de !important;
    color: #252523 !important;
    font-size: 0.88rem !important;
    font-weight: 600 !important;
    padding: 0.85rem 1rem !important;
    border-radius: 12px !important;
}

[data-testid="stExpander"] > div > div {
    background: #efe9de !important;
    padding: 0.5rem 1rem 1rem !important;
}

/* ── 13. Tabs ── */
[data-baseweb="tab-list"] {
    background: #efe9de !important;
    border-bottom: 1px solid #e6dfd8 !important;
    gap: 0 !important;
    padding: 0 0.5rem !important;
    border-radius: 12px 12px 0 0 !important;
}

[data-baseweb="tab"] {
    background: transparent !important;
    color: #6c6a64 !important;
    font-size: 0.84rem !important;
    font-weight: 500 !important;
    padding: 0.65rem 1.1rem !important;
    border-bottom: 2px solid transparent !important;
    transition: all 0.15s ease !important;
}

[data-baseweb="tab"]:hover {
    color: #252523 !important;
    background: #e8e0d2 !important;
}

[aria-selected="true"][data-baseweb="tab"] {
    color: #cc785c !important;
    border-bottom: 2px solid #cc785c !important;
    background: transparent !important;
    font-weight: 600 !important;
}

/* ── 14. DataFrames / Tables ── */
[data-testid="stDataFrame"],
[data-testid="stDataFrameResizable"] {
    border: 1px solid #e6dfd8 !important;
    border-radius: 12px !important;
    overflow: hidden !important;
}

.stDataFrame td,
.stDataFrame th,
[data-testid="stDataFrameResizable"] * {
    color: #252523 !important;
}

/* ── 15. File Uploader ── */
[data-testid="stFileUploaderDropzone"] {
    background: #efe9de !important;
    border: 2px dashed #e6dfd8 !important;
    border-radius: 12px !important;
    transition: all 0.2s ease !important;
}

[data-testid="stFileUploaderDropzone"]:hover {
    border-color: #cc785c !important;
    background: rgba(204,120,92,0.05) !important;
}

[data-testid="stFileUploaderDropzoneInstructions"] {
    text-align: right !important;
    color: #6c6a64 !important;
}

/* ── 16. Checkbox & Radio ── */
[data-testid="stCheckbox"] span,
[data-testid="stRadio"] span {
    color: #252523 !important;
    font-size: 0.84rem !important;
}

/* ── 17. Divider (---) ── */
hr {
    border: none !important;
    border-top: 1px solid #e6dfd8 !important;
    margin: 1.5rem 0 !important;
}

/* ── 18. Spinner ── */
[data-testid="stSpinner"] {
    color: #cc785c !important;
}

/* ── 19. Slider fix (LTR axis, RTL label) ── */
[data-testid="stSlider"] [data-baseweb="slider"] {
    direction: ltr !important;
}
[data-testid="stSlider"] [data-baseweb="slider"] * {
    direction: ltr !important;
}
[data-testid="stTickBarMin"], [data-testid="stTickBarMax"] {
    direction: ltr !important;
}

/* ── 20. Popover / Dropdown menus ── */
[data-baseweb="popover"],
[data-baseweb="menu"],
[role="listbox"] {
    background: #faf9f5 !important;
    border: 1px solid #e6dfd8 !important;
    border-radius: 8px !important;
    box-shadow: 0 8px 32px rgba(20,20,19,0.15) !important;
}

[data-baseweb="popover"] li,
[data-baseweb="menu"] li,
[role="option"] {
    direction: rtl !important;
    text-align: right !important;
    color: #252523 !important;
    font-size: 0.84rem !important;
}

[data-baseweb="menu"] li:hover,
[role="option"]:hover {
    background: #efe9de !important;
    color: #141413 !important;
}

/* ── 21. Float Apply Button ── */
.float-apply-bar {
    position: fixed;
    bottom: 24px;
    left: 50%;
    transform: translateX(-50%);
    z-index: 9999;
    background: #181715;
    border: 1.5px solid #5db872;
    border-radius: 50px;
    padding: 10px 28px;
    box-shadow: 0 4px 24px rgba(93,184,114,0.3), 0 0 0 1px rgba(93,184,114,0.08);
    display: flex;
    align-items: center;
    gap: 14px;
    animation: pulse-border 2.5s infinite;
    backdrop-filter: blur(12px);
}

@keyframes pulse-border {
    0%   { box-shadow: 0 4px 24px rgba(93,184,114,0.3); }
    50%  { box-shadow: 0 4px 40px rgba(93,184,114,0.55), 0 0 20px rgba(93,184,114,0.15); }
    100% { box-shadow: 0 4px 24px rgba(93,184,114,0.3); }
}

.float-apply-bar span {
    color: #5db872;
    font-size: 0.9rem;
    font-weight: 700;
    letter-spacing: 0.02em;
}

/* ── 22. Confirm Dialog overlay ── */
.confirm-overlay {
    position: fixed;
    inset: 0;
    background: rgba(20,20,19,0.55);
    z-index: 10000;
    display: flex;
    align-items: center;
    justify-content: center;
    backdrop-filter: blur(4px);
}

.confirm-box {
    background: #faf9f5;
    border: 1px solid #e6dfd8;
    border-radius: 14px;
    padding: 32px 36px;
    min-width: 340px;
    text-align: center;
    color: #141413;
    box-shadow: 0 16px 60px rgba(20,20,19,0.35);
}

/* ── 23. Charts / Bar charts ── */
[data-testid="stArrowVegaLiteChart"] {
    background: #efe9de !important;
    border: 1px solid #e6dfd8 !important;
    border-radius: 12px !important;
    padding: 1rem !important;
}

/* ── 24. Number input stepper buttons ── */
.stNumberInput button {
    background: #efe9de !important;
    border-color: #e6dfd8 !important;
    color: #6c6a64 !important;
    border-radius: 6px !important;
}

.stNumberInput button:hover {
    background: #e8e0d2 !important;
    color: #141413 !important;
}

/* ── 25. Font inheritance — force Arabic font everywhere ── */
*, *::before, *::after,
button, input, select, textarea,
.stMarkdown, .stMarkdown *,
[data-testid] * {
    font-family: 'IBM Plex Sans Arabic', 'StyreneB', 'Inter',
                 -apple-system, 'Segoe UI', system-ui, sans-serif !important;
}

/* ── 26. Metrics — uniform height & vertical alignment ── */
[data-testid="stMetric"] {
    display: flex !important;
    flex-direction: column !important;
    justify-content: center !important;
    min-height: 100px !important;
    height: 100% !important;
}

[data-testid="stMetricValue"] {
    font-size: 1.65rem !important;
    line-height: 1.2 !important;
    white-space: nowrap !important;
    overflow: hidden !important;
    text-overflow: ellipsis !important;
}

/* Equal height columns for metric rows */
[data-testid="stHorizontalBlock"] > div {
    display: flex !important;
    flex-direction: column !important;
}

/* ── 27. Day-edit Calendar Buttons — replace tiny icon with proper card ── */
/* The quick-edit buttons grid */
.day-btn-grid {
    display: grid;
    grid-template-columns: repeat(10, 1fr);
    gap: 6px;
    margin-top: 0.75rem;
    direction: rtl;
}

/* Style all buttons that are inside the day-edit section */
[data-testid="stHorizontalBlock"] .stButton > button {
    width: 100% !important;
    min-height: 52px !important;
    border-radius: 10px !important;
    font-size: 1.1rem !important;
    padding: 4px 2px !important;
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
    background: #faf9f5 !important;
    border: 1px solid #e6dfd8 !important;
    transition: all 0.13s ease !important;
    line-height: 1 !important;
}

[data-testid="stHorizontalBlock"] .stButton > button:hover {
    background: #efe9de !important;
    border-color: #cc785c !important;
    transform: translateY(-2px) !important;
    box-shadow: 0 4px 14px rgba(204,120,92,0.2) !important;
}

/* Caption above each day button (the number) */
[data-testid="stHorizontalBlock"] [data-testid="stCaptionContainer"] {
    text-align: center !important;
    color: #8e8b82 !important;
    font-size: 0.72rem !important;
    font-weight: 600 !important;
    letter-spacing: 0.04em !important;
    margin-bottom: 2px !important;
    display: block !important;
}

/* ── 28. Table — force visible text colors (override stDataFrame inline styles) ── */
[data-testid="stDataFrameResizable"] table {
    background: transparent !important;
}

[data-testid="stDataFrameResizable"] th {
    background: #efe9de !important;
    color: #6c6a64 !important;
    font-size: 0.78rem !important;
    font-weight: 600 !important;
    letter-spacing: 0.05em !important;
    text-transform: uppercase !important;
    border-bottom: 1px solid #e6dfd8 !important;
    padding: 0.65rem 0.85rem !important;
}

[data-testid="stDataFrameResizable"] td {
    font-size: 0.84rem !important;
    padding: 0.55rem 0.85rem !important;
    border-bottom: 1px solid rgba(230,223,216,0.7) !important;
}

/* Scrollbar inside dataframe */
[data-testid="stDataFrameResizable"] ::-webkit-scrollbar {
    width: 5px;
    height: 5px;
}
[data-testid="stDataFrameResizable"] ::-webkit-scrollbar-track { background: #faf9f5; }
[data-testid="stDataFrameResizable"] ::-webkit-scrollbar-thumb { background: #d9d2c5; border-radius: 3px; }

/* ── 29. Sidebar padding ── */
section[data-testid="stSidebar"] > div {
    padding-top: 1.5rem !important;
}

/* ── 30. File uploader text & icon visibility ── */
[data-testid="stFileUploaderDropzone"] p,
[data-testid="stFileUploaderDropzone"] span,
[data-testid="stFileUploaderDropzone"] small,
[data-testid="stFileUploaderDropzoneInstructions"] p,
[data-testid="stFileUploaderDropzoneInstructions"] span {
    color: #6c6a64 !important;
}

[data-testid="stFileUploaderDropzone"] svg {
    color: #cc785c !important;
    opacity: 0.8 !important;
}

/* Browse files button inside uploader */
[data-testid="stFileUploaderDropzone"] button {
    background: #efe9de !important;
    color: #252523 !important;
    border: 1px solid #e6dfd8 !important;
    border-radius: 6px !important;
    font-size: 0.82rem !important;
}

/* Fix uploader button text in new Streamlit versions */
[data-testid="stFileUploaderDropzone"] button p,
[data-testid="stFileUploaderDropzone"] button,
[data-testid="stFileUploaderDropzone"] button div {
    display: none !important;
}
[data-testid="stFileUploaderDropzone"] button::after {
    content: 'Browse files' !important;
    font-family: 'IBM Plex Sans Arabic', system-ui, sans-serif !important;
    font-size: 0.82rem !important;
}

/* ── 31. Select slider track and thumb ── */
[data-testid="stSlider"] [role="slider"] {
    background: #cc785c !important;
    border-color: #cc785c !important;
}

[data-testid="stSlider"] [data-baseweb="slider"] [data-testid="stTickBarMin"],
[data-testid="stSlider"] [data-baseweb="slider"] [data-testid="stTickBarMax"] {
    color: #8e8b82 !important;
}

/* ── 32. Global scrollbar ── */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #faf9f5; }
::-webkit-scrollbar-thumb { background: #e6dfd8; border-radius: 4px; }
::-webkit-scrollbar-thumb:hover { background: #d9d2c5; }

/* ── Prevent scroll jump on rerun ── */
html {
    scroll-behavior: auto !important;
}

/* Keep scroll position — Streamlit rerun anchor fix */
[data-testid="stAppViewContainer"] {
    overflow-anchor: auto !important;
}
</style>
""", unsafe_allow_html=True)


# ─── Helpers ──────────────────────────────────────────────────────────────────

def remove_close_punches(times_list: list, tolerance_minutes: int = 10) -> list:
    """
    حذف البصمات المتقاربة جداً من بعضها.

    مثال:
        times = ["8:00", "8:05", "8:20"]
        result = remove_close_punches(times, tolerance_minutes=10)
        # النتيجة: ["8:00", "8:20"]  (حذفت 8:05 لأنها أقل من 10 دقائق)
    """
    if not times_list or len(times_list) <= 1:
        return times_list

    filtered = [times_list[0]]  # الأولى دائماً تُحفظ

    for current_time_str in times_list[1:]:
        try:
            last_time = filtered[-1]
            last_h, last_m = map(int, last_time.split(':'))
            curr_h, curr_m = map(int, current_time_str.split(':'))

            last_total_min = last_h * 60 + last_m
            curr_total_min = curr_h * 60 + curr_m

            diff = abs(curr_total_min - last_total_min)

            if diff > tolerance_minutes:
                filtered.append(current_time_str)
        except (ValueError, IndexError):
            filtered.append(current_time_str)

    return filtered


def to_minutes(t):
    t = str(t).strip()
    if not t:
        return None
    parts = t.split(':')
    try:
        return int(parts[0]) * 60 + int(parts[1])
    except (ValueError, IndexError):
        return None


def minutes_diff(t_in, t_out):
    m_in  = to_minutes(t_in)
    m_out = to_minutes(t_out)
    if m_in is None or m_out is None:
        return None
    diff = m_out - m_in
    if diff < 0:
        diff += 24 * 60
    return diff


def fmt_hm(minutes):
    if minutes is None or minutes == 0:
        return "0:00"
    hh = int(minutes) // 60
    mm = int(minutes) % 60
    return f"{hh}:{mm:02d}"


def fmt_h(hours):
    if hours is None or hours == 0:
        return "0:00"
    return fmt_hm(round(hours * 60))


def fmt_days(n):
    if n is None:
        return "0"
    n = float(n)
    if n.is_integer():
        return str(int(n))
    return f"{n:.1f}"


def to_12h(t24: str) -> str:
    m = to_minutes(t24)
    if m is None:
        return t24
    hh = (m // 60) % 24
    mm = m % 60
    period = "م" if hh >= 12 else "ص"
    h12    = hh % 12 or 12
    return f"{h12}:{mm:02d} {period}"


def from_12h(h: int, mn: int, period: str) -> str:
    h = int(h); mn = int(mn)
    if period == "ص":
        hh24 = 0 if h == 12 else h
    else:
        hh24 = 12 if h == 12 else h + 12
    return f"{hh24:02d}:{mn:02d}"


def fmt_pair_12h(pair_str: str) -> str:
    corrected = pair_str.endswith("✓")
    clean = pair_str.rstrip("✓").strip()
    parts = clean.split("→")
    if len(parts) == 2:
        result = f"{to_12h(parts[0].strip())} ← {to_12h(parts[1].strip())}"
        return result + " ✓" if corrected else result
    return pair_str


# لون/سُمك البصمة المُصحَّحة يدويًا — لتمييزها بصريًا عن البصمة الأصلية المسجَّلة من الجهاز
CORRECTED_PUNCH_STYLE = "font-weight:700;color:#a9583e;"


def fmt_pair_12h_html(pair_str: str, corrected_side: str | None = None) -> str:
    """
    نفس fmt_pair_12h لكنه يرجّع HTML، ويجعل الطرف "المُصحَّح" (corrected_side = 'first' أو
    'second') بولد بخط عريض ولون مختلف عن الطرف "الأصلي" القادم فعليًا من بصمة الجهاز،
    عشان المستخدم يفرّق بينهم بصريًا بنظرة واحدة.
    """
    corrected = pair_str.endswith("✓")
    clean = pair_str.rstrip("✓").strip()
    parts = clean.split("→")
    if len(parts) != 2:
        return pair_str
    a = to_12h(parts[0].strip())
    b = to_12h(parts[1].strip())
    if corrected_side == 'first':
        a = f"<b style='{CORRECTED_PUNCH_STYLE}'>{a}</b>"
    elif corrected_side == 'second':
        b = f"<b style='{CORRECTED_PUNCH_STYLE}'>{b}</b>"
    result = f"{a} ← {b}"
    return result + " ✓" if corrected else result


def circular_mean(minutes_list: list) -> int:
    """
    حساب المتوسط الدائري لقائمة أوقات (بالدقائق) بحيث يعمل بشكل صحيح
    مع بصمات ما بعد منتصف الليل (مثلاً 23:00 و 01:00 متوسطهما 00:00 لا 12:00).

    الخوارزمية: تحويل كل وقت إلى زاوية على دائرة 360° ثم حساب متوسط المتجهات.
    """
    if not minutes_list:
        return 0
    n = len(minutes_list)
    # تحويل كل دقيقة إلى زاوية راديان على الدائرة (1440 دقيقة = 2π)
    sin_sum = sum(math.sin(2 * math.pi * m / 1440) for m in minutes_list)
    cos_sum = sum(math.cos(2 * math.pi * m / 1440) for m in minutes_list)
    angle   = math.atan2(sin_sum / n, cos_sum / n)
    # تحويل الزاوية مرة أخرى إلى دقائق [0, 1440)
    result  = round(angle * 1440 / (2 * math.pi)) % 1440
    return int(result)


def circular_dist(a: int, b: int) -> int:
    """المسافة الدائرية بين وقتين بالدقائق (تأخذ بعين الاعتبار تجاوز منتصف الليل)."""
    d = abs(a - b)
    return min(d, 1440 - d)


# ════════════════════════ نظام الأنماط (Shift Patterns) ════════════════════════
def cluster_times_circular(times_in_minutes, min_gap_minutes=120):
    """
    يقسم قائمة أوقات (بالدقائق) إلى مجموعات تمثل أنماط دوام مختلفة.
    يستخدم خوارزمية بسيطة: يرتب الأوقات دائرياً، ثم يبحث عن فجوات > min_gap_minutes.
    يعيد قائمة من المجموعات، كل مجموعة قائمة أوقات (بالدقائق).
    """
    if not times_in_minutes:
        return []
    sorted_times = sorted(times_in_minutes)
    n = len(sorted_times)
    if n <= 1:
        return [sorted_times]

    gaps = []
    for i in range(n):
        t1 = sorted_times[i]
        t2 = sorted_times[(i+1)%n]
        diff = (t2 - t1) % 1440
        gaps.append((diff, i))

    start_indices = []
    _, max_gap_idx = max(gaps, key=lambda x: x[0])
    current_idx = (max_gap_idx + 1) % n
    cluster = []
    for _ in range(n):
        t = sorted_times[current_idx]
        if cluster:
            last_t = cluster[-1]
            gap = (t - last_t) % 1440
            if gap >= min_gap_minutes:
                start_indices.append(current_idx)
        cluster.append(t)
        current_idx = (current_idx + 1) % n
        if current_idx == (max_gap_idx + 1) % n:
            break

    if not start_indices:
        return [sorted_times]

    clusters = []
    start = 0
    for idx in start_indices:
        clusters.append(sorted_times[start:idx])
        start = idx
    clusters.append(sorted_times[start:])
    clusters = [c for c in clusters if c]
    return clusters


def assign_shift_patterns(emp_days):
    """
    يحلل أيام كل موظف ويعين نمط دوام لكل يوم بناءً على وقت الحضور.
    يضيف مفتاح 'shift_pattern' (int) لكل يوم في emp_days.
    """
    for eid, days in emp_days.items():
        in_times = []
        valid_days_indices = []
        for di, d in enumerate(days):
            if d['raw_times'] and len(d['raw_times']) >= 2:
                in_m = to_minutes(d['raw_times'][0])
                if in_m is not None:
                    in_times.append(in_m)
                    valid_days_indices.append(di)

        if not in_times:
            for d in days:
                d['shift_pattern'] = None
            continue

        clusters = cluster_times_circular(in_times, min_gap_minutes=120)
        pattern_centers = []
        for cl in clusters:
            center = circular_mean(cl)
            pattern_centers.append(center)

        for di, d in enumerate(days):
            times = d.get('raw_times', [])
            if not times:
                d['shift_pattern'] = None
                continue
            first_time_min = to_minutes(times[0])
            if first_time_min is None:
                d['shift_pattern'] = None
                continue

            min_dist = float('inf')
            assigned_pattern = None
            for p_idx, center in enumerate(pattern_centers):
                dist = circular_dist(first_time_min, center)
                if dist < min_dist:
                    min_dist = dist
                    assigned_pattern = p_idx
            d['shift_pattern'] = assigned_pattern


def classify_lone_punch(lone_time: str, all_days: list, shift_pattern=None) -> dict:
    """
    يحدد هل البصمة الوحيدة في اليوم هي حضور أم انصراف.
    إذا تم توفير shift_pattern (رقم)، يتم فقط استخدام الأيام التي تنتمي لنفس النمط.

    Returns dict:
      'type': 'check_in' | 'check_out' | 'unknown'
      'confidence': 'high' | 'low'
      'avg_in': str  (متوسط وقت الحضور بصيغة HH:MM)
      'avg_out': str (متوسط وقت الانصراف بصيغة HH:MM)
      'dist_to_in': int  (المسافة بالدقائق للحضور)
      'dist_to_out': int (المسافة بالدقائق للانصراف)
      'corr_suggestion': str | None  (اقتراح مرتبط بالوقت المعروف)
      'corr_sample_size': int  (عدد الأيام المستخدمة في الاقتراح المرتبط)
    """
    lone_min = to_minutes(lone_time)
    if lone_min is None:
        return {'type': 'unknown', 'confidence': 'low', 'avg_in': None, 'avg_out': None,
                'dist_to_in': None, 'dist_to_out': None,
                'corr_suggestion': None, 'corr_sample_size': 0}

    # تصفية الأيام بناءً على النمط إذا وجد
    if shift_pattern is not None:
        relevant_days = [d for d in all_days if d.get('shift_pattern') == shift_pattern]
    else:
        relevant_days = all_days

    check_in_mins  = []
    check_out_mins = []
    pairs_list     = []   # قائمة (in_min, out_min) من الأيام المكتملة

    for d in relevant_days:
        times = d.get('raw_times', [])
        # نأخذ فقط الأيام التي فيها بصمتان أو أكثر لنستخرج الحضور والانصراف المؤكدين
        if len(times) >= 2:
            first_min = to_minutes(times[0])
            last_min  = to_minutes(times[-1])
            if first_min is not None:
                check_in_mins.append(first_min)
            if last_min is not None:
                check_out_mins.append(last_min)
            for i in range(0, len(times) - 1, 2):
                in_m  = to_minutes(times[i])
                out_m = to_minutes(times[i + 1])
                if in_m is not None and out_m is not None:
                    raw_diff = out_m - in_m
                    if raw_diff < 0:
                        raw_diff += 1440
                    if 2 <= raw_diff <= 960:
                        pairs_list.append((in_m, out_m))

    if not check_in_mins or not check_out_mins:
        # إذا لم توجد بيانات كافية في هذا النمط، نوسع لجميع الأيام (fallback)
        if shift_pattern is not None:
            return classify_lone_punch(lone_time, all_days, shift_pattern=None)
        return {'type': 'unknown', 'confidence': 'low', 'avg_in': None, 'avg_out': None,
                'dist_to_in': None, 'dist_to_out': None,
                'corr_suggestion': None, 'corr_sample_size': 0}

    # ── المتوسط الدائري: يعمل صح مع منتصف الليل ──────────────────────────
    avg_in  = circular_mean(check_in_mins)
    avg_out = circular_mean(check_out_mins)

    dist_in  = circular_dist(lone_min, avg_in)
    dist_out = circular_dist(lone_min, avg_out)

    avg_in_str  = fmt_hm(avg_in)
    avg_out_str = fmt_hm(avg_out)

    if dist_in < dist_out:
        punch_type = 'check_in'
        confidence = 'high' if (dist_out - dist_in) > 30 else 'low'
    elif dist_out < dist_in:
        punch_type = 'check_out'
        confidence = 'high' if (dist_in - dist_out) > 30 else 'low'
    else:
        punch_type = 'unknown'
        confidence = 'low'

    # ── الاقتراح المرتبط بالوقت المعروف ──────────────────────────────────
    WINDOW = 45
    corr_suggestion    = None
    corr_sample_size   = 0

    if punch_type == 'check_in' and pairs_list:
        matching_outs = [
            out for (inp, out) in pairs_list
            if circular_dist(inp, lone_min) <= WINDOW
        ]
        if matching_outs:
            corr_suggestion  = fmt_hm(circular_mean(matching_outs))
            corr_sample_size = len(matching_outs)

    elif punch_type == 'check_out' and pairs_list:
        matching_ins = [
            inp for (inp, out) in pairs_list
            if circular_dist(out, lone_min) <= WINDOW
        ]
        if matching_ins:
            corr_suggestion  = fmt_hm(circular_mean(matching_ins))
            corr_sample_size = len(matching_ins)

    return {
        'type':             punch_type,
        'confidence':       confidence,
        'avg_in':           avg_in_str,
        'avg_out':          avg_out_str,
        'dist_to_in':       dist_in,
        'dist_to_out':      dist_out,
        'corr_suggestion':  corr_suggestion,
        'corr_sample_size': corr_sample_size,
    }


def time_input_12h(label: str, key: str, existing_24h: str = "", color: str = "") -> str | None:
    def_h, def_m, def_p = 8, 0, "ص"
    if existing_24h:
        mins = to_minutes(existing_24h)
        if mins is not None:
            hh = (mins // 60) % 24
            mm = mins % 60
            def_p = "م" if hh >= 12 else "ص"
            def_h = hh % 12 or 12
            def_m = mm

    if color:
        st.markdown(
            f"<div style='background:#efe9de;border-radius:8px;padding:6px 10px 2px 10px;"
            f"margin-bottom:4px'><small>{label}</small></div>",
            unsafe_allow_html=True
        )
    else:
        st.caption(label)

    c1, c2, c3 = st.columns([2, 2, 2])
    with c1:
        hour = st.number_input(
            "ساعة", min_value=1, max_value=12,
            value=def_h, step=1,
            key=f"{key}_h", label_visibility="collapsed"
        )
        st.caption("ساعة")
    with c2:
        minute = st.number_input(
            "دقيقة", min_value=0, max_value=59,
            value=def_m, step=1,
            key=f"{key}_m", label_visibility="collapsed"
        )
        st.caption("دقيقة")
    with c3:
        period = st.selectbox(
            "ص/م", options=["ص", "م"],
            index=0 if def_p == "ص" else 1,
            key=f"{key}_p", label_visibility="collapsed"
        )
        st.caption("ص / م")

    return from_12h(hour, minute, period)


# ─── Save/Load corrections ────────────────────────────────────────────────────

CORRECTIONS_SHEET = "corrections_data"


def load_corrections_from_excel(file_bytes):
    corrections   = {}
    hourly_rates  = {}
    day_overrides = {}
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        if CORRECTIONS_SHEET not in wb.sheetnames:
            return corrections, hourly_rates, day_overrides
        ws = wb[CORRECTIONS_SHEET]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            key, sub_key, value, rtype = row
            if rtype == "correction":
                if key not in corrections:
                    corrections[key] = {}
                corrections[key][sub_key] = value
            elif rtype == "rate":
                try:
                    hourly_rates[key] = float(value)
                except (ValueError, TypeError):
                    pass
            elif rtype == "day_override":
                if key not in day_overrides:
                    day_overrides[key] = {}
                if sub_key == "status":
                    day_overrides[key]['status'] = value
                elif sub_key == "pairs":
                    try:
                        day_overrides[key]['pairs'] = [tuple(p) for p in json.loads(value)]
                    except (ValueError, TypeError, json.JSONDecodeError):
                        day_overrides[key]['pairs'] = []
    except Exception as e:
        st.warning(f"تعذّر قراءة ملف التصحيحات: {e}")
    return corrections, hourly_rates, day_overrides


def get_file_hash(file_bytes):
    return hashlib.md5(file_bytes).hexdigest()


# نافذة "ما بعد منتصف الليل" المشتركة (من 12:00 ص إلى 6:00 ص) — بالدقائق.
# يُستخدم نفس الحد الأقصى لكل من: مدى slider الـ Redistribute (cutoff)
# ومدى slider الـ Saturate، ولتحديد نطاق عمل دالة saturate_punches.
# مهم: القيمتان (cutoff و saturate) مستقلتان تمامًا عن بعضهما — هذا الثابت
# فقط يحدد سقف المدى المسموح لكل واحدة منهما على حدة.
OVERNIGHT_WINDOW_MAX = 360  # 6:00 ص


def redistribute_overnight_punches(raw_punches: dict, cutoff_hour: float = 3.0) -> dict:
    cutoff_min = int(cutoff_hour * 60)
    result = {eid: {day: list(times) for day, times in days.items()}
              for eid, days in raw_punches.items()}

    for eid, days in result.items():
        sorted_days = sorted(days.keys())
        for i, day in enumerate(sorted_days):
            times = days[day]
            if not times:
                continue

            overnight = []
            remaining = []
            for t in times:
                m = to_minutes(t)
                if m is not None and m < cutoff_min:
                    overnight.append(t)
                else:
                    remaining.append(t)

            if not overnight:
                continue

            if i > 0:
                prev_day = sorted_days[i - 1]
                days[prev_day] = days[prev_day] + overnight
            days[day] = remaining

    return result


def saturate_punches(raw_punches: dict, saturate_min: int,
                      window_max: int = OVERNIGHT_WINDOW_MAX) -> dict:
    """
    دالة مستقلة تمامًا عن حد الـ Redistribute (cutoff) — وتُطبَّق فقط على
    آخر بصمة في قائمة اليوم (الانصراف)، وليس على بصمة الحضور (الأولى)
    أو أي بصمة وسطى — حتى لو وقعت هذه البصمات داخل نافذة ما بعد منتصف
    الليل بالصدفة (مثل حضور الساعة 5:55 ص في شيفت صباحي عادي).

    إذا كانت آخر بصمة في اليوم تقع داخل نافذة ما بعد منتصف الليل (من 0
    إلى window_max دقيقة، أي حتى 6:00 ص) وتتجاوز saturate_min دقيقة،
    تُستبدل بـ saturate_min نفسه (Clamp/Saturate). أي بصمة أخرى في
    اليوم — أولى أو وسطى — لا تُلمس أبدًا بغض النظر عن قيمتها.

    مثال 1: يوم فيه [05:55 (حضور), 14:25 (انصراف)]، saturate_min = 60
        - آخر بصمة (الانصراف) = 14:25 → خارج النافذة (>360) → بدون تغيير
        - بصمة الحضور 05:55 لا تُفحص أبدًا، فتبقى كما هي

    مثال 2: يوم فيه [14:00 (حضور), 00:52 (انصراف بعد منتصف الليل)]، saturate_min = 50
        - آخر بصمة (الانصراف) = 00:52 (52 د) → داخل النافذة وتتجاوز 50 → تصبح 00:50
    """
    sh  = saturate_min // 60
    sm  = saturate_min % 60
    cap = f"{sh:02d}:{sm:02d}"

    result = {}
    for eid, days in raw_punches.items():
        result[eid] = {}
        for day, times in days.items():
            if not times:
                result[eid][day] = times
                continue

            new_times = list(times)
            m = to_minutes(new_times[-1])   # آخر بصمة فقط = الانصراف
            if m is not None and saturate_min < m <= window_max:
                new_times[-1] = cap

            result[eid][day] = new_times
    return result


def _build_summary_df(emp_days, log_names, log_depts):
    """دالة مشتركة: تحوّل emp_days إلى DataFrame ملخّص للموظفين."""
    rows = []
    for eid, days in emp_days.items():
        total_min      = sum(d['total_min'] for d in days)
        att_days_cnt   = sum(1 for d in days if d['status'] == 'حضور' and d['total_min'] > 0)
        absent_cnt     = sum(1 for d in days if d['status'] == 'غياب')
        incomplete_cnt = sum(1 for d in days if d['incomplete'])
        rows.append({
            'id':              eid,
            'name':            log_names.get(eid, eid),
            'department':      log_depts.get(eid, ''),
            'work_minutes':    total_min,
            'work_hours':      round(total_min / 60, 2),
            'attendance_days': att_days_cnt,
            'absent_days':     absent_cnt,
            'incomplete_days': incomplete_cnt,
        })
    return pd.DataFrame(rows)


@st.cache_data(show_spinner=False)
def parse_file_hikvision(file_bytes, cutoff_hour: float = 3.0, saturate_min: int = None, duplicate_punch_tolerance: int = 10):
    """
    يقرأ ملف HikVision بتنسيق AttendanceRecord.
    الهيكل:
      - ورقة واحدة اسمها 'AttendanceRecord'
      - صف 3 (index): Employee ID | Card No. | Name | Department | 2025/06/01 | 2025/06/02 | ...
      - صف 4: '' | '' | '' | '' | SW - EW | SW - EW | ...
      - صفوف 5+: كل صف = موظف واحد
      - كل خلية يوم = 4 أسطر مفصولة بـ \\n، كل سطر = "HH:MM HH:MM"
        حيث --:-- = لا يوجد بصمة
    يضمن:
      - أي عدد أيام (28/29/30/31 أو نص شهر) ← ديناميكي من headers
      - بصمة في موضع الحضور فقط  → needs_review
      - بصمة في موضع الانصراف فقط (--:-- HH:MM) → needs_review
      - كلا البصمتين موجودتان → زوج مكتمل
    """
    wb = xlrd.open_workbook(file_contents=file_bytes)
    sh = wb.sheet_by_name('AttendanceRecord')

    # ── بناء خريطة عمود → رقم اليوم (ديناميكي) ──────────────────────────
    col_to_day = {}
    for c in range(sh.ncols):
        header = str(sh.cell_value(3, c)).strip()   # مثال: '2025/06/15'
        if '/' in header:
            try:
                day_num = int(header.split('/')[-1])  # آخر جزء = اليوم
                col_to_day[c] = day_num
            except (ValueError, IndexError):
                pass

    raw_punches = {}
    log_names   = {}
    log_depts   = {}

    # ── قراءة بيانات الموظفين (من صف 5 وما بعده) ─────────────────────────
    for r in range(5, sh.nrows):
        eid  = str(sh.cell_value(r, 0)).strip()
        name = str(sh.cell_value(r, 2)).strip()
        dept = str(sh.cell_value(r, 3)).strip()

        if not eid or not name:
            continue
        # تحويل eid لرقم صحيح إن أمكن (يزيل .0 من الأرقام)
        try:
            eid = str(int(float(eid)))
        except (ValueError, TypeError):
            pass

        log_names[eid] = name
        log_depts[eid] = dept

        punches = {}
        for c, day_num in col_to_day.items():
            cell_val = str(sh.cell_value(r, c)).strip()
            if not cell_val:
                continue

            # كل سطر = جلسة واحدة: "HH:MM HH:MM" أو "--:-- HH:MM" إلخ
            day_times = []
            for line in cell_val.split('\n'):
                parts = line.strip().split()
                for p in parts:
                    if re.match(r'^\d{1,2}:\d{2}$', p):
                        day_times.append(('real', p))
                    elif p == '--:--':
                        day_times.append(('empty', p))

            # نعالج الجلسات: كل سطر = زوج (حضور, انصراف)
            real_times = []
            i = 0
            while i + 1 < len(day_times):
                t_in_type,  t_in  = day_times[i]
                t_out_type, t_out = day_times[i + 1]
                i += 2

                has_in  = (t_in_type  == 'real')
                has_out = (t_out_type == 'real')

                if has_in and has_out:
                    real_times.extend([t_in, t_out])
                elif has_in and not has_out:
                    real_times.append(t_in)
                elif not has_in and has_out:
                    real_times.append(t_out)
                # كلاهما فارغ → تجاهل

            if real_times:
                # حذف البصمات المتقاربة
                real_times = remove_close_punches(real_times, tolerance_minutes=duplicate_punch_tolerance)
                punches[day_num] = real_times

        raw_punches[eid] = punches

    # ── ضمان وجود كل أيام الشهر قبل Redistribute ──────────────────────────
    # بدون ده، اليوم الأول الفارغ مش موجود في الـ dict فالـ redistribute
    # بيشوفه كأول يوم وما يلاقيش سابقه فيوقف الترحيل
    _all_known_days = set(col_to_day.values())
    for _eid in raw_punches:
        for _day in _all_known_days:
            raw_punches[_eid].setdefault(_day, [])

    # ── تطبيق Redistribute ────────────────────────────────────────────────
    raw_punches = redistribute_overnight_punches(raw_punches, cutoff_hour=cutoff_hour)

    # ── تطبيق Saturate ────────────────────────────────────────────────────
    if saturate_min is not None and saturate_min > 0:
        raw_punches = saturate_punches(raw_punches, saturate_min=saturate_min)

    # ── بناء emp_days (نفس هيكل parse_file القديم تماماً) ─────────────────
    # HikVision لا يوفر معلومات الغياب/العطل في الملف → نعتمد على البصمات فقط
    emp_days = {}

    for eid, punches in raw_punches.items():
        # كل أيام الشهر المصدَّرة (من headers) — حتى لو الموظف ما جاش
        all_days = set(col_to_day.values()) | set(punches.keys())
        days = []

        for day_num in sorted(all_days):
            times = punches.get(day_num, [])

            pairs = []
            for i in range(0, len(times) - 1, 2):
                diff = minutes_diff(times[i], times[i + 1])
                if diff is not None:
                    pairs.append((f"{times[i]}→{times[i+1]}", diff))

            has_unpaired = (len(times) % 2 == 1)
            unpaired_time = times[-1] if has_unpaired else None

            total_min  = sum(m for _, m in pairs)
            incomplete = has_unpaired

            needs_review = []
            if has_unpaired:
                needs_review.append({
                    'type':        'missing_out',
                    'ci':          unpaired_time,
                    'co':          None,
                    'resolved':    False,
                    'resolved_co': None,
                    'suggested':   None,
                })

            # HikVision: لو فيه بصمات = حضور، لو مفيش = غياب
            # (الملف لا يميّز بين الغياب والعطلة)
            if times:
                status = 'حضور'
            else:
                status = 'غياب'

            days.append({
                'day':          f"{day_num:02d}",
                'status':       status,
                'total_min':    total_min,
                'incomplete':   incomplete,
                'punch_pairs':  pairs,
                'needs_review': needs_review,
                'raw_times':    times,
                'tolerance_added': 0,
            })

        emp_days[eid] = days

    # تعيين أنماط الدوام
    assign_shift_patterns(emp_days)

    return _build_summary_df(emp_days, log_names, log_depts), emp_days


@st.cache_data(show_spinner=False)
def parse_file(file_bytes, cutoff_hour: float = 3.0, saturate_min: int = None, duplicate_punch_tolerance: int = 10):
    wb = xlrd.open_workbook(file_contents=file_bytes)

    log_sh = wb.sheet_by_name('Att.log report')

    col_to_day = {}
    for c in range(log_sh.ncols):
        v = str(log_sh.cell_value(3, c)).strip()
        if v:
            try:
                col_to_day[c] = int(float(v))
            except (ValueError, TypeError):
                pass

    raw_punches = {}
    log_names   = {}
    log_depts   = {}

    for r in range(log_sh.nrows):
        if str(log_sh.cell_value(r, 0)).strip() != 'ID:':
            continue
        eid  = str(log_sh.cell_value(r, 2)).strip()
        name = str(log_sh.cell_value(r, 10)).strip()
        dept = str(log_sh.cell_value(r, 20)).strip()
        log_names[eid] = name
        log_depts[eid] = dept

        punch_row = r + 1
        if punch_row >= log_sh.nrows:
            continue
        punches = {}
        for c, day in col_to_day.items():
            if c >= log_sh.ncols:
                continue
            v = str(log_sh.cell_value(punch_row, c)).strip()
            if v:
                times = re.findall(r'\d{1,2}:\d{2}', v)
                if times:
                    # حذف البصمات المتقاربة جداً
                    times = remove_close_punches(times, tolerance_minutes=duplicate_punch_tolerance)
                    punches[day] = times
        raw_punches[eid] = punches

    # ── ضمان وجود كل أيام الشهر قبل Redistribute ──────────────────────────
    _all_known_days = set(col_to_day.values())
    for _eid in raw_punches:
        for _day in _all_known_days:
            raw_punches[_eid].setdefault(_day, [])

    raw_punches = redistribute_overnight_punches(raw_punches, cutoff_hour=cutoff_hour)

    # ── تطبيق Saturate بعد Redistribute ──────────────────────────────────
    if saturate_min is not None and saturate_min > 0:
        raw_punches = saturate_punches(raw_punches, saturate_min=saturate_min)

    skip = {'Schedule Infor.', 'Att. Stat.', 'Att.log report', 'Exception Stat.'}
    absent_info = {}

    for sname in wb.sheet_names():
        if sname in skip:
            continue
        sh = wb.sheet_by_name(sname)
        for offset in [0, 15, 30]:
            name_col = offset + 9
            if name_col >= sh.ncols:
                continue
            eid = str(sh.cell_value(3, name_col)).strip()
            if not eid or eid == 'ID':
                continue
            day_status = {}
            for r in range(11, min(41, sh.nrows)):
                day_label = str(sh.cell_value(r, offset + 0)).strip()
                if not day_label:
                    continue
                m = re.match(r'^(\d+)', day_label)
                if not m:
                    continue
                day_num = int(m.group(1))
                on1 = str(sh.cell_value(r, offset + 1)).strip()
                has_any = any(
                    str(sh.cell_value(r, offset + c)).strip()
                    for c in [1, 3, 6, 8, 10, 12]
                    if offset + c < sh.ncols
                )
                if on1 == 'Absent':
                    day_status[day_num] = 'absent'
                elif not has_any:
                    day_status[day_num] = 'weekend'
            absent_info[eid] = day_status

    emp_days = {}

    for eid, punches in raw_punches.items():
        all_days = set(absent_info.get(eid, {}).keys()) | set(punches.keys())
        days = []

        for day_num in sorted(all_days):
            status_marker = absent_info.get(eid, {}).get(day_num, 'present')
            times = punches.get(day_num, [])

            pairs = []
            for i in range(0, len(times) - 1, 2):
                diff = minutes_diff(times[i], times[i + 1])
                if diff is not None:
                    pairs.append((f"{times[i]}→{times[i+1]}", diff))

            has_unpaired = (len(times) % 2 == 1)
            unpaired_time = times[-1] if has_unpaired else None

            total_min  = sum(m for _, m in pairs)
            incomplete = has_unpaired

            needs_review = []
            if has_unpaired:
                needs_review.append({
                    'type':        'missing_out',
                    'ci':          unpaired_time,
                    'co':          None,
                    'resolved':    False,
                    'resolved_co': None,
                    'suggested':   None,
                })

            if status_marker == 'absent' or (status_marker == 'weekend' and not times):
                status = 'غياب'
            else:
                status = 'حضور'

            days.append({
                'day':          f"{day_num:02d}",
                'status':       status,
                'total_min':    total_min,
                'incomplete':   incomplete,
                'punch_pairs':  pairs,
                'needs_review': needs_review,
                'raw_times':    times,
                'tolerance_added': 0,
            })

        emp_days[eid] = days

    # تعيين أنماط الدوام
    assign_shift_patterns(emp_days)

    return _build_summary_df(emp_days, log_names, log_depts), emp_days


# ════════════════════════ نظام التسامح مع المغادرة المبكرة ════════════════════════
def apply_early_tolerance(emp_days, tolerance_minutes, tolerance_enabled, saturate_min=None):
    """
    تطبيق التعويض العادل للمغادرة المبكرة باستخدام حد الـ Saturate كمرجع ثابت.
    
    المنطق الجديد:
    - إذا تم تفعيل التسامح (tolerance_enabled) وكان saturate_min > 0،
    - نأخذ وقت الانصراف الفعلي (آخر بصمة في اليوم) ونتحقق مما إذا كان يقع في النافذة:
        [saturate_min - tolerance_minutes, saturate_min) (مع مراعاة الدورة اليومية 1440 دقيقة).
    - إذا تحقق الشرط، نرفع وقت الانصراف إلى saturate_time ونضيف الفرق إلى إجمالي ساعات اليوم.
    - يتم تسجيل الفرق المضاف في حقل tolerance_added لكل يوم.
    - لا نعتمد على أي متوسطات مستنتجة من البيانات، بل نستخدم saturate_min كمعيار ثابت.
    """
    if not tolerance_enabled or tolerance_minutes <= 0 or saturate_min is None or saturate_min <= 0:
        return emp_days

    # تحويل saturate_min إلى نص HH:MM
    sh = saturate_min // 60
    sm = saturate_min % 60
    saturate_time = f"{sh:02d}:{sm:02d}"

    # نطاق التسامح بالدقائق
    tolerance = tolerance_minutes

    for eid, days in emp_days.items():
        for d in days:
            # نطبق فقط على الأيام التي حالتها حضور ولديها بصمات (على الأقل 2)
            if d['status'] != 'حضور':
                continue
            raw_times = d.get('raw_times', [])
            if len(raw_times) < 2:
                continue

            # آخر بصمة = وقت الانصراف الفعلي
            actual_out_str = raw_times[-1]
            actual_out = to_minutes(actual_out_str)
            if actual_out is None:
                continue

            # نحسب الفرق بين saturate_min و actual_out مع مراعاة الدورة اليومية
            # الفرق الإيجابي يعني أن actual_out قبل saturate_min (أي انصرف مبكراً)
            diff = (saturate_min - actual_out) % 1440
            # إذا كان الفرق أكبر من 720 (نصف اليوم)، فهذا يعني أن actual_out بعد saturate_min
            # (قد يحدث إذا كان saturate_min صغيراً مثلاً 30 دقيقة، وactual_out = 23:30 أي 1410 دقيقة)
            # في هذه الحالة لا نريد تطبيق التعويض لأن الانصراف متأخر جداً
            if diff > 720:
                continue

            # نتحقق مما إذا كان الفرق ضمن نطاق التسامح (0 < diff <= tolerance)
            if 0 < diff <= tolerance:
                # سيتم رفع وقت الانصراف إلى saturate_time
                # نحسب الفرق المضاف (بالدقائق) لتحديث total_min
                added_minutes = diff

                # تحديث total_min
                d['total_min'] += added_minutes

                # تسجيل الفرق المضاف في حقل tolerance_added
                d['tolerance_added'] = added_minutes

                # تحديث raw_times (آخر عنصر)
                d['raw_times'][-1] = saturate_time

                # تحديث punch_pairs أيضاً لأنها مستخدمة في العرض
                # نحتاج إلى تحديث الزوج الأخير في punch_pairs إن وجد
                if d['punch_pairs']:
                    # نأخذ آخر زوج ونعدل وقت الانصراف فيه
                    last_pair = d['punch_pairs'][-1]
                    pair_str, old_diff = last_pair
                    # نستبدل الوقت القديم بالجديد
                    parts = pair_str.rstrip("✓").strip().split('→')
                    if len(parts) == 2:
                        # parts[0] = الحضور, parts[1] = الانصراف القديم
                        suffix = "✓" if pair_str.endswith("✓") else ""
                        new_pair_str = f"{parts[0]}→{saturate_time}{suffix}"
                        # الفرق الجديد = الفرق القديم + الدقائق المضافة
                        # (أسلم من إعادة حساب minutes_diff لتجنب مشاكل عبور منتصف الليل)
                        new_diff = old_diff + added_minutes
                        d['punch_pairs'][-1] = (new_pair_str, new_diff)

    return emp_days


def apply_overrides(emp_days):
    # نسخ محلي حتى لا نُعدّل بيانات الـ session_state الأصلية بالخطأ
    corrections   = {k: dict(v) for k, v in st.session_state.get('corrections', {}).items()}
    day_overrides = dict(st.session_state.get('day_overrides', {}))

    # ── دمج التعديلات المعلَّقة (لم يُضغط بعد على "Apply") ──────────────
    # بدون هذا الدمج، جدول العرض اليومي (اللي بيعاين pending_changes) يُظهر
    # اليوم كأنه مُصحَّح ومحسوب، بينما الإجمالي الفعلي هنا (المستخدم في
    # الرواتب والمؤشرات العلوية) كان يتجاهل هذا التعديل تمامًا لأنه لم
    # يُطبَّق بعد. هذا كان يسبب فروقًا غير مفهومة بين الجدول والإجمالي.
    pending = st.session_state.get('pending_changes', {})
    for pv in pending.values():
        ptype = pv.get('type')
        if ptype == 'correction':
            kb = pv['key_base']
            corrections.setdefault(kb, {})
            corrections[kb][pv['rev_key']] = pv['value']
            corrections[kb][f"{pv['rev_key']}_role"] = pv.get('punch_role', 'check_in')
        elif ptype == 'day_override':
            day_overrides[pv['key_base']] = pv['data']
        elif ptype == 'remove_override':
            day_overrides.pop(pv['key_base'], None)

    # الحصول على إعدادات التسامح من session_state
    tol_enabled = st.session_state.get('_tolerance_enabled', False)
    tol_minutes = st.session_state.get('_tolerance_minutes', 0)

    # تطبيق التسامح أولاً على الأيام الأصلية (قبل التعديلات اليدوية)
    saturate_min = st.session_state.get('_saturate_minutes', None)
    emp_days = apply_early_tolerance(emp_days, tol_minutes, tol_enabled, saturate_min)

    summary = {}

    for eid, days in emp_days.items():
        total_min       = 0
        att_days_cnt    = 0.0
        absent_days_cnt = 0
        incomplete_cnt  = 0

        for di, d in enumerate(days):
            key_base = f"{eid}_{di}"

            if key_base in day_overrides:
                ov_day = day_overrides[key_base]
                status = ov_day.get('status')
                pairs  = ov_day.get('pairs') or []

                day_total = 0
                for ci, co in pairs:
                    diff = minutes_diff(ci, co)
                    if diff is not None:
                        day_total += diff
                total_min += day_total

                if status == 'حضور':
                    att_days_cnt += 1
                elif status == 'غياب':
                    absent_days_cnt += 1
                continue

            day_corrections = corrections.get(key_base, {})

            base_min  = sum(m for _, m in d['punch_pairs'])
            extra_min = 0
            still_inc = False

            for ri, rev in enumerate(d.get('needs_review', [])):
                rev_key      = f"{key_base}_r{ri}"
                resolved_time = day_corrections.get(rev_key)
                if resolved_time:
                    punch_role = day_corrections.get(f"{rev_key}_role", 'check_in')
                    if punch_role == 'check_out':
                        # البصمة المعروفة انصراف، المُصحَّحة حضور
                        diff = minutes_diff(resolved_time, rev['ci'])
                    else:
                        # البصمة المعروفة حضور، المُصحَّحة انصراف (الافتراضي)
                        diff = minutes_diff(rev['ci'], resolved_time)
                    if diff is not None:
                        extra_min += diff
                else:
                    still_inc = True

            day_total  = base_min + extra_min
            total_min += day_total

            # ── تحديد الحالة الفعلية لليوم ──────────────────────────────
            # لو اليوم أصلاً "غياب" في الملف الخام (لأن فيه بصمة واحدة
            # لونية فقط) لكن تم تصحيحه بإضافة البصمة الناقصة وأصبح له
            # ساعات عمل فعلية، لازم يُحتسب "حضور" فعليًا - وليس "غياب"
            # كما كان يحدث سابقًا بالاعتماد على d['status'] الأصلي فقط.
            effective_status = d['status']
            if effective_status == 'غياب' and day_total > 0:
                effective_status = 'حضور'

            if effective_status == 'حضور' and (day_total > 0 or not still_inc):
                att_days_cnt += 1
            if effective_status == 'غياب':
                absent_days_cnt += 1
            if still_inc and d['incomplete']:
                incomplete_cnt += 1

        summary[eid] = {
            'work_minutes':    total_min,
            'work_hours':      round(total_min / 60, 2),
            'attendance_days': att_days_cnt,
            'absent_days':     absent_days_cnt,
            'incomplete_days': incomplete_cnt,
        }

    return summary


# ─── Payroll ──────────────────────────────────────────────────────────────────

def calculate_payroll(df, hourly_rates, overrides_summary):
    rows = []
    for _, emp in df.iterrows():
        eid  = str(emp['id'])
        rate = hourly_rates.get(eid, 0.0)
        if rate <= 0:
            continue
        ov          = overrides_summary.get(eid, {})
        work_h      = ov.get('work_hours',      emp.get('work_hours', 0))
        att_days    = ov.get('attendance_days', emp.get('attendance_days', 0))
        absent_days = ov.get('absent_days',     emp.get('absent_days', 0))
        inc_days    = ov.get('incomplete_days', emp.get('incomplete_days', 0))
        net = round(work_h * rate, 2)
        rows.append({
            'ID':                  eid,
            'الاسم':               emp['name'],
            'القسم':               emp.get('department', ''),
            'أيام الحضور':         round(float(att_days), 1),
            'أيام الغياب':         int(absent_days),
            'أيام بصمة ناقصة':     int(inc_days),
            'ساعات العمل الفعلية': round(work_h, 2),
            'سعر الساعة':          rate,
            'صافي الراتب':         net,
        })
    return pd.DataFrame(rows)


# ─── Export ───────────────────────────────────────────────────────────────────

def export_to_excel(payroll_df):
    wb2 = Workbook()
    ws  = wb2.active
    ws.title = "كشف الرواتب"
    hfill = PatternFill("solid", start_color="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    alt   = PatternFill("solid", start_color="D9E1F2")
    thin  = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'),  bottom=Side(style='thin'))
    headers = list(payroll_df.columns)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = hfont; cell.fill = hfill
        cell.alignment = Alignment(horizontal='center'); cell.border = thin
    for i, row in enumerate(payroll_df.itertuples(index=False), 2):
        ws.append(list(row))
        f = alt if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        for cell in ws[i]:
            cell.fill = f; cell.border = thin
            cell.alignment = Alignment(horizontal='center')
    lr = ws.max_row + 1
    ws.cell(lr, 1, "الإجمالي").font = Font(bold=True, name="Arial")
    for ci in [7, 9]:
        ws.cell(lr, ci, f"=SUM({get_column_letter(ci)}2:{get_column_letter(ci)}{lr-1})")
        ws.cell(lr, ci).font = Font(bold=True, name="Arial")
    for col in ws.columns:
        mx = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(12, min(mx + 4, 30))
    buf = io.BytesIO()
    wb2.save(buf)
    buf.seek(0)
    return buf.read()


def export_full_file(payroll_df, corrections, hourly_rates, day_overrides=None):
    day_overrides = day_overrides or {}
    wb2 = Workbook()

    ws = wb2.active
    ws.title = "كشف الرواتب"
    hfill = PatternFill("solid", start_color="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    alt   = PatternFill("solid", start_color="D9E1F2")
    thin  = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'),  bottom=Side(style='thin'))
    headers = list(payroll_df.columns) if not payroll_df.empty else []
    if headers:
        ws.append(headers)
        for cell in ws[1]:
            cell.font = hfont; cell.fill = hfill
            cell.alignment = Alignment(horizontal='center'); cell.border = thin
        for i, row in enumerate(payroll_df.itertuples(index=False), 2):
            ws.append(list(row))
            f = alt if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
            for cell in ws[i]:
                cell.fill = f; cell.border = thin
                cell.alignment = Alignment(horizontal='center')
        lr = ws.max_row + 1
        ws.cell(lr, 1, "الإجمالي").font = Font(bold=True, name="Arial")
        for ci in [7, 9]:
            ws.cell(lr, ci, f"=SUM({get_column_letter(ci)}2:{get_column_letter(ci)}{lr-1})")
            ws.cell(lr, ci).font = Font(bold=True, name="Arial")
        for col in ws.columns:
            mx = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = max(12, min(mx + 4, 30))

    ws2 = wb2.create_sheet(CORRECTIONS_SHEET)
    ws2.append(["key", "sub_key", "value", "type"])
    for key_base, sub_dict in corrections.items():
        for sub_key, value in sub_dict.items():
            ws2.append([key_base, sub_key, value, "correction"])
    for eid, rate in hourly_rates.items():
        ws2.append([eid, "hourly_rate", str(rate), "rate"])
    for key_base, ov in day_overrides.items():
        ws2.append([key_base, "status", ov.get('status', ''), "day_override"])
        ws2.append([key_base, "pairs", json.dumps(ov.get('pairs', [])), "day_override"])

    buf = io.BytesIO()
    wb2.save(buf)
    buf.seek(0)
    return buf.read()


# ─── Pending Changes Tracker ──────────────────────────────────────────────────
# نتتبع التعديلات المعلّقة (لم تُطبَّق بعد)

def count_pending_changes():
    """إرجاع عدد التعديلات المعلّقة التي تنتظر الـ Apply"""
    pending = st.session_state.get('pending_changes', {})
    return len(pending)


def add_pending_change(change_key: str, change_data: dict):
    """إضافة تعديل للقائمة المعلّقة"""
    if 'pending_changes' not in st.session_state:
        st.session_state.pending_changes = {}
    st.session_state.pending_changes[change_key] = change_data


def apply_all_pending():
    """تطبيق جميع التعديلات المعلّقة على البيانات الفعلية"""
    pending = st.session_state.get('pending_changes', {})
    if not pending:
        return 0

    applied = 0
    for change_key, change_data in pending.items():
        ctype = change_data.get('type')

        if ctype == 'correction':
            # تصحيح بصمة ناقصة
            key_base = change_data['key_base']
            rev_key  = change_data['rev_key']
            value    = change_data['value']
            punch_role = change_data.get('punch_role', 'check_in')
            if key_base not in st.session_state.corrections:
                st.session_state.corrections[key_base] = {}
            st.session_state.corrections[key_base][rev_key] = value
            st.session_state.corrections[key_base][f"{rev_key}_role"] = punch_role
            applied += 1

        elif ctype == 'day_override':
            # تعديل يوم بالكامل
            key_base = change_data['key_base']
            status   = change_data['status']
            pairs    = change_data['pairs']
            st.session_state.day_overrides[key_base] = {'status': status, 'pairs': pairs}
            applied += 1

        elif ctype == 'remove_override':
            # حذف تعديل يوم
            key_base = change_data['key_base']
            if key_base in st.session_state.day_overrides:
                del st.session_state.day_overrides[key_base]
            applied += 1

    st.session_state.pending_changes = {}
    
    # ✅ حفظ الحالة الجديدة تلقائياً بعد التطبيق
    auto_save_file_state()
    
    return applied


def discard_all_pending():
    """تجاهل جميع التعديلات المعلّقة"""
    st.session_state.pending_changes = {}


# ─── Float Apply Bar ──────────────────────────────────────────────────────────

def show_float_apply_bar():
    """عرض شريط Apply العائم إذا كانت هناك تعديلات معلّقة"""
    n = count_pending_changes()
    if n == 0:
        return

    label = "✅ Apply" if n == 1 else f"✅ Apply all ({n} تعديلات)"

    # ── Ctrl+Enter shortcut ───────────────────────────────────────────────
    st.markdown("""
    <script>
    (function() {
        if (window._applyHooked) return;
        window._applyHooked = true;
        document.addEventListener('keydown', function(e) {
            if ((e.ctrlKey || e.metaKey) && e.key === 'Enter') {
                e.preventDefault();
                var btns = Array.from(document.querySelectorAll('button'));
                var applyBtn = btns.find(function(b) {
                    return b.innerText && b.innerText.includes('Apply');
                });
                if (applyBtn) { applyBtn.click(); }
            }
        });
    })();
    </script>
    """, unsafe_allow_html=True)

    st.markdown("---")
    apply_col, discard_col, hint_col = st.columns([2, 2, 4])
    with apply_col:
        if st.button(label, type="primary", key="float_apply_btn", use_container_width=True):
            count = apply_all_pending()
            st.success(f"✅ تم تطبيق {count} تعديل بنجاح!")
            st.rerun()
    with discard_col:
        if st.button("🗑️ تجاهل التعديلات", key="float_discard_btn", use_container_width=True):
            discard_all_pending()
            st.rerun()
    with hint_col:
        st.caption("💡 اختصار: Ctrl + Enter")


# ─── Unsaved Changes Warning Dialog ──────────────────────────────────────────

def check_unsaved_before_leave(destination_label: str = "") -> bool:
    """
    يتحقق إذا كان هناك تعديلات معلّقة قبل مغادرة الصفحة.
    يعرض Popup تأكيد.
    يُرجع True إذا يمكن المغادرة، False إذا يجب البقاء.
    """
    n = count_pending_changes()
    if n == 0:
        return True  # لا يوجد شيء معلّق، يمكن المغادرة

    # يوجد تعديلات معلّقة — نعرض Dialog
    dialog_key = "unsaved_dialog_open"
    if dialog_key not in st.session_state:
        st.session_state[dialog_key] = False

    st.session_state[dialog_key] = True

    if st.session_state.get(dialog_key):
        st.warning(f"⚠️ لديك {n} تعديل لم يُطبَّق بعد! هل تريد حفظه قبل المغادرة؟")
        col1, col2, col3 = st.columns(3)
        with col1:
            if st.button("💾 حفظ وانتقال", key="unsaved_save_leave", type="primary"):
                apply_all_pending()
                st.session_state[dialog_key] = False
                return True
        with col2:
            if st.button("❌ إلغاء (ابقَ في الصفحة)", key="unsaved_cancel"):
                st.session_state[dialog_key] = False
                return False
        with col3:
            if st.button("🗑️ تجاهل وانتقال", key="unsaved_discard_leave"):
                discard_all_pending()
                st.session_state[dialog_key] = False
                return True
        return False  # لا نغادر حتى يختار

    return True


# ─── Review panel (with pending support) ─────────────────────────────────────

def show_review_panel(eid, days):
    """
    عرض البصمات الناقصة.
    - إذا تمت مراجعة البصمة وتأكيدها (Apply) تختفي من هنا.
    - التعديل يُضاف للـ pending حتى يضغط Apply.
    """
    corrections  = st.session_state.get('corrections', {})
    pending      = st.session_state.get('pending_changes', {})

    # نحسب البصمات التي لم تُحلَّ بعد (لا في corrections ولا في pending)
    actually_pending = []
    for di, d in enumerate(days):
        if not d.get('needs_review'):
            continue
        key_base = f"{eid}_{di}"
        if key_base in st.session_state.get('day_overrides', {}):
            continue
        for ri, rev in enumerate(d['needs_review']):
            rev_key = f"{key_base}_r{ri}"
            if corrections.get(key_base, {}).get(rev_key):
                continue
            pending_key = f"corr_{key_base}_{ri}"
            if pending_key in pending:
                continue
            actually_pending.append((di, d, ri, rev, key_base, rev_key))

    if not actually_pending:
        st.success("✅ لا توجد بصمات تحتاج مراجعة")
        return

    # ── حساب الاقتراحات الذكية المتاحة لكل بصمة (مسبقاً) ────────────────
    smart_info = {}   # rev_key → {'suggestion': str, 'punch_type': str, 'ci': str}
    for (di, d, ri, rev, key_base, rev_key) in actually_pending:
        ci  = rev.get('ci', '')
        sp  = d.get('shift_pattern')  # نمط اليوم الحالي
        clf = classify_lone_punch(ci, days, shift_pattern=sp)
        smart_info[rev_key] = {
            'clf':       clf,
            'ci':        ci,
            'key_base':  key_base,
            'ri':        ri,
            'd':         d,
        }

    has_smart = any(
        si['clf']['corr_suggestion'] and si['clf']['type'] in ('check_in', 'check_out')
        for si in smart_info.values()
    )

    # ── أزرار العمليات الجماعية في صفحة الموظف ──────────────────────────
    smart_count = sum(
        1 for si in smart_info.values()
        if si['clf']['corr_suggestion']
        and si['clf']['type'] in ('check_in', 'check_out')
        and si['clf']['confidence'] == 'high'
    )
    bulk_col1, bulk_col2 = st.columns(2)

    with bulk_col1:
        if st.button(
            "✅ تأكيد كل التعديلات",
            key=f"bulk_confirm_{eid}",
            type="primary",
            use_container_width=True
        ):
            confirmed = 0
            for rev_key, si in smart_info.items():
                clf        = si['clf']
                ci         = si['ci']
                key_base   = si['key_base']
                ri         = si['ri']
                d          = si['d']
                punch_type = clf['type']

                # ما الخيار المحدد حالياً في هذه البصمة؟
                current_sel = st.session_state.get(f"sel_{rev_key}", "لم يُحدَّد بعد")

                # النوع الفعلي مع مراعاة التبديل اليدوي
                flip_state   = st.session_state.get(f"flip_type_{rev_key}", False)
                eff_type     = punch_type
                if flip_state and punch_type == 'check_in':
                    eff_type = 'check_out'
                elif flip_state and punch_type == 'check_out':
                    eff_type = 'check_in'

                chosen      = None
                final_role  = eff_type

                if current_sel.startswith("✨") and clf['corr_suggestion']:
                    chosen     = clf['corr_suggestion']
                    final_role = eff_type
                elif current_sel == "تحديد يدوي":
                    # نقرأ قيمة الـ time_input_12h من session_state
                    h_key = f"man_{rev_key}_h"
                    m_key = f"man_{rev_key}_m"
                    p_key = f"man_{rev_key}_p"
                    if h_key in st.session_state:
                        chosen = from_12h(
                            st.session_state[h_key],
                            st.session_state[m_key],
                            st.session_state[p_key]
                        )
                        final_role = eff_type

                if chosen:
                    pending_key = f"corr_{key_base}_{ri}"
                    add_pending_change(pending_key, {
                        'type':       'correction',
                        'key_base':   key_base,
                        'rev_key':    rev_key,
                        'value':      chosen,
                        'punch_role': final_role,
                    })
                    confirmed += 1

            if confirmed:
                st.success(f"✅ تمت إضافة {confirmed} تصحيح للقائمة المعلّقة — اضغط Apply لتطبيقها")
                st.rerun()
            else:
                st.warning("⚠️ لم يتم تحديد أي وقت بعد — اختر الاقتراح الذكي أو حدد يدوياً أولاً")

    with bulk_col2:
        if has_smart and smart_count > 0:
            if st.button(
                f"🤖 تطبيق الاقتراح الذكي ({smart_count} بصمة)",
                key=f"bulk_smart_emp_{eid}",
                type="secondary",
                use_container_width=True
            ):
                applied_smart = 0
                for rev_key, si in smart_info.items():
                    clf = si['clf']
                    if (clf['type'] in ('check_in', 'check_out')
                            and clf['confidence'] == 'high'
                            and clf['corr_suggestion']):
                        pending_key = f"corr_{si['key_base']}_{si['ri']}"
                        add_pending_change(pending_key, {
                            'type':       'correction',
                            'key_base':   si['key_base'],
                            'rev_key':    rev_key,
                            'value':      clf['corr_suggestion'],
                            'punch_role': clf['type'],
                        })
                        applied_smart += 1
                if applied_smart:
                    st.success(f"✅ تمت إضافة {applied_smart} تصحيح ذكي للقائمة المعلّقة")
                    st.rerun()
                else:
                    st.warning("⚠️ لا توجد بصمات بثقة عالية 🟢 قابلة للتطبيق")
        else:
            st.button(
                "🤖 تطبيق الاقتراح الذكي",
                key=f"bulk_smart_emp_{eid}",
                disabled=True,
                use_container_width=True
            )

    st.divider()

    # ── عرض البصمات فردياً ───────────────────────────────────────────────
    for (di, d, ri, rev, key_base, rev_key) in actually_pending:
        st.markdown(f"**يوم {d['day']}**")

        si         = smart_info[rev_key]
        ci         = si['ci']
        clf        = si['clf']
        punch_type       = clf['type']
        confidence       = clf['confidence']
        avg_in_str       = clf['avg_in']
        avg_out_str      = clf['avg_out']
        dist_in          = clf['dist_to_in']
        dist_out         = clf['dist_to_out']
        corr_suggestion  = clf['corr_suggestion']
        corr_sample_size = clf['corr_sample_size']

        if d['punch_pairs']:
            pairs_str = "  |  ".join(
                f"{fmt_pair_12h(p[0])} ({fmt_hm(p[1])})" for p in d['punch_pairs']
            )
            st.caption(f"الفترات المكتملة: {pairs_str}")

        # ── مفتاح التبديل اليدوي لنوع البصمة ────────────────────────────
        flip_key        = f"flip_type_{rev_key}"
        flip_confirmed  = f"flip_confirmed_{rev_key}"
        
        # ✅ تحميل نوع البصمة المحفوظ إذا كان موجوداً
        saved_fingerprint_type = None
        saved_types = st.session_state.get('fingerprint_types', {})
        if eid in saved_types and rev_key in saved_types[eid]:
            saved_fingerprint_type = saved_types[eid][rev_key]
        
        # إذا كان هناك نوع محفوظ، طبّقه مباشرة
        if saved_fingerprint_type:
            flipped = (saved_fingerprint_type != punch_type)
        else:
            flipped = st.session_state.get(flip_key, False)

        # النوع الفعلي المستخدم (قد يكون مقلوباً بواسطة المستخدم)
        effective_type = punch_type
        if flipped and punch_type == 'check_in':
            effective_type = 'check_out'
        elif flipped and punch_type == 'check_out':
            effective_type = 'check_in'

        # إعادة حساب اقتراح الاقتران بناءً على النوع الفعلي
        if effective_type != punch_type:
            # نعيد حساب corr_suggestion بالنوع المقلوب
            pairs_list_flip = []
            for fd in days:
                ftimes = fd.get('raw_times', [])
                if len(ftimes) >= 2:
                    for i in range(0, len(ftimes) - 1, 2):
                        in_m  = to_minutes(ftimes[i])
                        out_m = to_minutes(ftimes[i + 1])
                        if in_m is not None and out_m is not None:
                            raw_diff = out_m - in_m
                            if raw_diff < 0:
                                raw_diff += 1440
                            if 2 <= raw_diff <= 960:
                                pairs_list_flip.append((in_m, out_m))

            lone_min_val = to_minutes(ci)
            WINDOW_FLIP  = 45
            if effective_type == 'check_in' and pairs_list_flip and lone_min_val is not None:
                matching_outs = [o for (inp, o) in pairs_list_flip if circular_dist(inp, lone_min_val) <= WINDOW_FLIP]
                if matching_outs:
                    corr_suggestion  = fmt_hm(circular_mean(matching_outs))
                    corr_sample_size = len(matching_outs)
                else:
                    corr_suggestion = None; corr_sample_size = 0
            elif effective_type == 'check_out' and pairs_list_flip and lone_min_val is not None:
                matching_ins = [inp for (inp, o) in pairs_list_flip if circular_dist(o, lone_min_val) <= WINDOW_FLIP]
                if matching_ins:
                    corr_suggestion  = fmt_hm(circular_mean(matching_ins))
                    corr_sample_size = len(matching_ins)
                else:
                    corr_suggestion = None; corr_sample_size = 0

        if effective_type == 'check_in':
            conf_label = "🟢 ثقة عالية" if confidence == 'high' else "🟡 ثقة متوسطة"
            if flipped:
                warn_msg = (
                    f"🔄 بصمة {to_12h(ci)} — **تم تغييرها يدوياً إلى: حضور**\n\n"
                    f"(كان البرنامج يقترح انصراف — تم تجاوزه)"
                )
            else:
                warn_msg = (
                    f"⚠️ بصمة {to_12h(ci)} — بصمة **حضور** على الأرجح {conf_label}\n\n"
                    f"متوسط حضور الموظف: {to_12h(avg_in_str)} (فرق {dist_in} د) | "
                    f"متوسط انصرافه: {to_12h(avg_out_str)} (فرق {dist_out} د) — **لا يوجد انصراف مقابل لها**"
                )
            st.warning(warn_msg)
            if corr_suggestion and corr_sample_size > 0:
                st.info(
                    f"💡 **اقتراح ذكي:** في الأيام التي حضر فيها الموظف قرب {to_12h(ci)}، "
                    f"كان ينصرف عادةً **{to_12h(corr_suggestion)}** "
                    f"(من {corr_sample_size} يوم مشابه)"
                )
            missing_label = "أدخل وقت الانصراف الناقص"
            sel_label     = "وقت الانصراف"

            # زر التبديل — لا يظهر إلا إذا كان البرنامج اقترح نوعاً محدداً
            if punch_type in ('check_in', 'check_out'):
                flip_btn_label = "🔄 لا، هي بصمة انصراف — غيّر نوعها" if not flipped else "↩️ إلغاء التبديل (إعادة لاقتراح البرنامج)"
                if st.button(flip_btn_label, key=f"flipbtn_{rev_key}"):
                    st.session_state[flip_key] = not flipped
                    # نمسح تأكيد التبديل القديم عند العدول
                    st.session_state.pop(flip_confirmed, None)
                    
                    # ✅ حفظ نوع البصمة المختار
                    eid = key_base.split('_')[0]
                    if eid not in st.session_state.fingerprint_types:
                        st.session_state.fingerprint_types[eid] = {}
                    st.session_state.fingerprint_types[eid][rev_key] = 'check_out' if not flipped else 'check_in'
                    auto_save_file_state()
                    
                    st.rerun()

        elif effective_type == 'check_out':
            conf_label = "🟢 ثقة عالية" if confidence == 'high' else "🟡 ثقة متوسطة"
            if flipped:
                warn_msg = (
                    f"🔄 بصمة {to_12h(ci)} — **تم تغييرها يدوياً إلى: انصراف**\n\n"
                    f"(كان البرنامج يقترح حضور — تم تجاوزه)"
                )
            else:
                warn_msg = (
                    f"⚠️ بصمة {to_12h(ci)} — بصمة **انصراف** على الأرجح {conf_label}\n\n"
                    f"متوسط انصراف الموظف: {to_12h(avg_out_str)} (فرق {dist_out} د) | "
                    f"متوسط حضوره: {to_12h(avg_in_str)} (فرق {dist_in} د) — **لا يوجد حضور مقابل لها**"
                )
            st.warning(warn_msg)
            if corr_suggestion and corr_sample_size > 0:
                st.info(
                    f"💡 **اقتراح ذكي:** في الأيام التي انصرف فيها الموظف قرب {to_12h(ci)}، "
                    f"كان يحضر عادةً **{to_12h(corr_suggestion)}** "
                    f"(من {corr_sample_size} يوم مشابه)"
                )
            missing_label = "أدخل وقت الحضور الناقص"
            sel_label     = "وقت الحضور"

            # زر التبديل
            if punch_type in ('check_in', 'check_out'):
                flip_btn_label = "🔄 لا، هي بصمة حضور — غيّر نوعها" if not flipped else "↩️ إلغاء التبديل (إعادة لاقتراح البرنامج)"
                if st.button(flip_btn_label, key=f"flipbtn_{rev_key}"):
                    st.session_state[flip_key] = not flipped
                    st.session_state.pop(flip_confirmed, None)
                    
                    # ✅ حفظ نوع البصمة المختار
                    eid = key_base.split('_')[0]
                    if eid not in st.session_state.fingerprint_types:
                        st.session_state.fingerprint_types[eid] = {}
                    st.session_state.fingerprint_types[eid][rev_key] = 'check_in' if not flipped else 'check_out'
                    auto_save_file_state()
                    
                    st.rerun()

        else:
            st.warning(
                f"⚠️ بصمة {to_12h(ci)} — لم يتمكن البرنامج من تحديد نوعها "
                f"(بيانات غير كافية للمقارنة) — تحتاج مراجعة يدوية"
            )
            missing_label = "أدخل الوقت الناقص"
            sel_label     = "الوقت الناقص"

        # ── خيارات الإدخال ──────────────────────────────────────────────
        options = ["لم يُحدَّد بعد"]
        smart_option_label = None
        if corr_suggestion and effective_type in ('check_in', 'check_out'):
            smart_option_label = f"✨ استخدام الاقتراح الذكي ({to_12h(corr_suggestion)})"
            options.append(smart_option_label)
        options.append("تحديد يدوي")

        # إذا كانت القيمة المخزَّنة سابقاً (مثلاً اقتراح ذكي بنص/وقت قديم قبل
        # التبديل) لم تعد ضمن الخيارات الحالية، نعيدها للافتراضي.
        # مهم: هذا التحقق يقارن بالخيارات الفعلية الحالية فقط — وليس شرطاً
        # غير مشروط كلما كانت flipped=True — لأن ذلك كان يمسح اختيار
        # "✨ الاقتراح الذكي" الصحيح في نفس اللحظة التي يضغط فيها المستخدم
        # زر "تأكيد التصحيح" (لأن الزر نفسه يسبب rerun وflipped لا تزال True
        # حتى تلك اللحظة)، فيرجع الاختيار إلى "لم يُحدَّد بعد" بدل تنفيذ التأكيد.
        sel_key = f"sel_{rev_key}"
        if st.session_state.get(sel_key) not in options:
            st.session_state[sel_key] = "لم يُحدَّد بعد"

        col_a, col_b = st.columns([2, 3])
        with col_a:
            sel = st.selectbox(sel_label, options=options, key=f"sel_{rev_key}")
        with col_b:
            if sel == "تحديد يدوي":
                chosen = time_input_12h(
                    missing_label,
                    key=f"man_{rev_key}",
                    existing_24h=""
                )
            elif sel.startswith("✨") and corr_suggestion:
                chosen = corr_suggestion
                if effective_type == 'check_out':
                    diff = minutes_diff(chosen, ci)
                    st.info(f"🕐 {to_12h(chosen)} (حضور) ← {to_12h(ci)} (انصراف) = {fmt_hm(diff)}")
                else:
                    diff = minutes_diff(ci, chosen)
                    st.info(f"🕐 {to_12h(ci)} (حضور) ← {to_12h(chosen)} (انصراف) = {fmt_hm(diff)}")
                pending_key = f"corr_{key_base}_{ri}"
                if st.button(f"✅ تأكيد تصحيح يوم {d['day']}", key=f"confirm_corr_{rev_key}", type="primary"):
                    add_pending_change(pending_key, {
                        'type':       'correction',
                        'key_base':   key_base,
                        'rev_key':    rev_key,
                        'value':      chosen,
                        'punch_role': effective_type,
                    })
                    # مسح حالة التبديل بعد التأكيد
                    st.session_state.pop(flip_key, None)
                    st.toast(f"✅ تم تسجيل تصحيح يوم {d['day']} — اضغط Apply لتطبيقه نهائياً", icon="✅")
                    st.rerun()
            else:
                chosen = None

        if chosen is None or not sel.startswith("✨"):
            if chosen:
                if effective_type == 'check_out':
                    diff = minutes_diff(chosen, ci)
                    st.info(f"🕐 {to_12h(chosen)} (حضور) ← {to_12h(ci)} (انصراف) = {fmt_hm(diff)}")
                else:
                    diff = minutes_diff(ci, chosen)
                    st.info(f"🕐 {to_12h(ci)} (حضور) ← {to_12h(chosen)} (انصراف) = {fmt_hm(diff)}")

                pending_key = f"corr_{key_base}_{ri}"
                if st.button(f"✅ تأكيد تصحيح يوم {d['day']}", key=f"confirm_corr_{rev_key}", type="primary"):
                    add_pending_change(pending_key, {
                        'type':       'correction',
                        'key_base':   key_base,
                        'rev_key':    rev_key,
                        'value':      chosen,
                        'punch_role': effective_type,
                    })
                    # مسح حالة التبديل بعد التأكيد
                    st.session_state.pop(flip_key, None)
                    st.toast(f"✅ تم تسجيل تصحيح يوم {d['day']} — اضغط Apply لتطبيقه نهائياً", icon="✅")
                    st.rerun()

        st.divider()


# ─── Day Editor (with pending support) ───────────────────────────────────────

DAY_STATUS_OPTIONS = ["حضور", "غياب"]
DAY_STATUS_TO_INTERNAL = {"حضور": "حضور", "غياب": "غياب"}
DAY_STATUS_FROM_INTERNAL = {v: k for k, v in DAY_STATUS_TO_INTERNAL.items()}
NO_CHANGE_LABEL = "بدون تعديل (استخدام بيانات البصمة الأصلية)"


def show_day_editor(eid, days, preselect_di=None):
    day_overrides = st.session_state.setdefault('day_overrides', {})
    pending       = st.session_state.get('pending_changes', {})

    # ── ملخص الأيام المعدَّلة (مؤكدة + معلّقة) ──
    emp_overrides = {k: v for k, v in day_overrides.items() if k.startswith(f"{eid}_")}
    pending_overrides = {
        k.replace(f"dayov_{eid}_", f"{eid}_"): v['data']
        for k, v in pending.items()
        if k.startswith(f"dayov_{eid}_") and v.get('type') == 'day_override'
    }
    all_overrides = {**emp_overrides, **pending_overrides}

    if all_overrides:
        st.markdown("**✏️ أيام معدَّلة لهذا الموظف:**")
        for key_base, ov in sorted(all_overrides.items(), key=lambda kv: int(kv[0].split('_')[-1])):
            di = int(key_base.split('_')[-1])
            day_label = days[di]['day'] if di < len(days) else key_base
            is_pending = key_base in {
                k.replace(f"dayov_{eid}_", f"{eid}_")
                for k in pending if k.startswith(f"dayov_{eid}_")
            }
            tag = " ⏳ (معلّق)" if is_pending else ""
            c_txt, c_del = st.columns([5, 1])
            with c_txt:
                hrs = sum((minutes_diff(a, b) or 0) for a, b in (ov.get('pairs') or []))
                st.caption(f"يوم {day_label} ← **{ov.get('status')}**"
                           + (f" ({fmt_hm(hrs)})" if hrs else "") + tag)
            with c_del:
                if key_base in emp_overrides and not is_pending:
                    if st.button("🗑️ إلغاء", key=f"quick_remove_{key_base}"):
                        pending_key = f"dayov_remove_{key_base}"
                        add_pending_change(pending_key, {
                            'type':     'remove_override',
                            'key_base': key_base,
                        })
                        st.rerun()
        st.divider()

    st.caption(
        "اختر اليوم الذي تريد تعديله: يمكنك تحويله بالكامل لحضور أو غياب أو نصف "
        "يوم أو إجازة، أو إعادة كتابة أوقات الدخول والانصراف الصحيحة."
    )

    day_labels = []
    day_lookup = {}
    for di, d in enumerate(days):
        key_base = f"{eid}_{di}"
        tag = "  ✏️ معدَّل" if key_base in all_overrides else ""
        day_labels.append(f"يوم {d['day']} — {d['status']}{tag}")
        day_lookup[day_labels[-1]] = di

    default_sel_idx = 0
    if preselect_di is not None:
        for idx, lbl in enumerate(day_labels):
            if day_lookup[lbl] == preselect_di:
                default_sel_idx = idx
                break

    sel_label = st.selectbox("اختر اليوم", day_labels, index=default_sel_idx,
                              key=f"day_sel_{eid}{'_q' if preselect_di is not None else ''}")
    di       = day_lookup[sel_label]
    d        = days[di]
    key_base = f"{eid}_{di}"
    existing = all_overrides.get(key_base)

    raw_times = d.get('raw_times', [])
    if raw_times:
        raw_str = "  ،  ".join(to_12h(t) for t in raw_times)
        st.info(f"📍 البصمات الخام ({len(raw_times)}): {raw_str}")
    else:
        st.info("📍 لا توجد أي بصمة مسجَّلة لهذا اليوم في الجهاز.")

    if d['punch_pairs']:
        cur_str = "  |  ".join(f"{fmt_pair_12h(p[0])} ({fmt_hm(p[1])})" for p in d['punch_pairs'])
        st.caption(f"الحساب التلقائي الحالي: {cur_str}")

    default_label = NO_CHANGE_LABEL
    if existing and existing.get('status') in DAY_STATUS_FROM_INTERNAL:
        default_label = DAY_STATUS_FROM_INTERNAL[existing['status']]

    options = [NO_CHANGE_LABEL] + DAY_STATUS_OPTIONS
    chosen_label = st.selectbox(
        "الحالة الجديدة لليوم بعد التعديل",
        options, index=options.index(default_label),
        key=f"day_status_{key_base}"
    )

    if chosen_label == NO_CHANGE_LABEL:
        if existing and st.button("🗑️ إلغاء التعديل والعودة للحساب التلقائي", key=f"remove_ov_{key_base}"):
            pending_key = f"dayov_remove_{key_base}"
            add_pending_change(pending_key, {
                'type':     'remove_override',
                'key_base': key_base,
            })
            st.info("⏳ سيُطبَّق الإلغاء عند الضغط على Apply")
        return

    new_status = DAY_STATUS_TO_INTERNAL[chosen_label]
    pairs_out  = []

    if new_status == "حضور":
        if existing and existing.get('pairs'):
            init_pairs = existing['pairs']
        else:
            init_pairs = []
            for pstr, _ in d.get('punch_pairs', []):
                parts = pstr.rstrip("✓").strip().split("→")
                if len(parts) == 2:
                    init_pairs.append((parts[0].strip(), parts[1].strip()))
                else:
                    init_pairs.append((parts[0].strip(), ""))
            for rev in d.get('needs_review', []):
                if rev.get('type') == 'missing_out' and rev.get('ci'):
                    init_pairs.append((rev['ci'], ""))
            if not init_pairs:
                init_pairs = [("", "")]

        default_n = len(init_pairs)
        n_pairs = st.number_input(
            "عدد فترات الحضور في هذا اليوم",
            min_value=1, max_value=6, value=max(1, default_n), step=1,
            key=f"day_npairs_{key_base}"
        )

        next_day_times = []
        if di + 1 < len(days):
            nd = days[di + 1]
            for t in nd.get('raw_times', []):
                next_day_times.append(t)

        for pi in range(int(n_pairs)):
            exist_in  = init_pairs[pi][0] if pi < len(init_pairs) else ""
            exist_out = init_pairs[pi][1] if pi < len(init_pairs) else ""

            st.markdown(
                f"<div style='background:#efe9de;border-radius:8px;padding:4px 12px;"
                f"margin:6px 0 2px 0'><b>الفترة {pi + 1}</b></div>",
                unsafe_allow_html=True
            )

            c_in, c_out = st.columns(2)
            with c_in:
                t_in = time_input_12h(
                    "🟢 وقت الدخول",
                    key=f"day_in_{key_base}_{pi}",
                    existing_24h=exist_in,
                    color="#1a3d1a"
                )
            with c_out:
                borrow_key = f"borrow_{key_base}_{pi}"
                borrowed   = st.session_state.get(borrow_key)

                if borrowed:
                    t_out = time_input_12h(
                        f"🔴 وقت الانصراف  ← مأخوذ من اليوم التالي ({to_12h(borrowed)})",
                        key=f"day_out_{key_base}_{pi}",
                        existing_24h=borrowed,
                        color="#3d1a1a"
                    )
                else:
                    t_out = time_input_12h(
                        "🔴 وقت الانصراف",
                        key=f"day_out_{key_base}_{pi}",
                        existing_24h=exist_out,
                        color="#3d1a1a"
                    )

            if next_day_times:
                opts = ["— لا شيء —"] + [f"{to_12h(t)}  ({t})" for t in next_day_times]
                cur_idx = 0
                if borrowed:
                    for idx, t in enumerate(next_day_times, 1):
                        if t == borrowed:
                            cur_idx = idx
                            break
                sel = st.selectbox(
                    f"↩️ سحب انصراف الفترة {pi+1} من بصمات اليوم التالي",
                    options=opts,
                    index=cur_idx,
                    key=f"borrow_sel_{key_base}_{pi}"
                )
                if sel == "— لا شيء —":
                    st.session_state[borrow_key] = None
                else:
                    chosen_t = next_day_times[opts.index(sel) - 1]
                    st.session_state[borrow_key] = chosen_t

            pairs_out.append((t_in, t_out))
            if pi < int(n_pairs) - 1:
                st.divider()

        preview_min = sum((minutes_diff(a, b) or 0) for a, b in pairs_out)
        st.success(f"⏱️ إجمالي ساعات اليوم بعد التعديل: {fmt_hm(preview_min)}")
    else:
        st.caption("هذا اليوم سيُحتسب غياباً ولن تُحسب له أي ساعات.")

    # ── أزرار الحفظ ──────────────────────────────────────────────────────
    btn_col1, btn_col2 = st.columns(2)

    with btn_col1:
        if st.button("✅ تأكيد تعديل اليوم", key=f"confirm_day_{key_base}", type="primary"):
            pending_key = f"dayov_{eid}_{di}"
            add_pending_change(pending_key, {
                'type':     'day_override',
                'key_base': key_base,
                'status':   new_status,
                'pairs':    pairs_out,
                'data':     {'status': new_status, 'pairs': pairs_out},
            })
            apply_all_pending()
            st.success(f"✅ تم تطبيق تعديل يوم {d['day']} بنجاح!")
            st.rerun()

    with btn_col2:
        if st.button("➕ إضافة للقائمة المعلّقة فقط", key=f"queue_day_{key_base}", type="secondary"):
            pending_key = f"dayov_{eid}_{di}"
            add_pending_change(pending_key, {
                'type':     'day_override',
                'key_base': key_base,
                'status':   new_status,
                'pairs':    pairs_out,
                'data':     {'status': new_status, 'pairs': pairs_out},
            })

            st.rerun()


# ─── Employee detail view ──────────────────────────────────────────────────────

def show_employee_detail(emp_row, emp_days, overrides_summary):
    eid  = str(emp_row['id'])
    name = emp_row['name']
    days = emp_days.get(eid, [])

    # ── Ctrl+Enter يشغّل زر Apply ────────────────────────────────────────
    st.markdown("""
    <script>
    (function() {
        function hookCtrlEnter() {
            document.addEventListener('keydown', function(e) {
                if ((e.ctrlKey || e.metaKey) && e.key === 'Enter') {
                    e.preventDefault();
                    // نبحث عن زر Apply (float_apply_btn أو أي زر primary يحتوي Apply)
                    var btns = Array.from(document.querySelectorAll('button[kind="primary"], button'));
                    var applyBtn = btns.find(function(b) {
                        return b.innerText && b.innerText.includes('Apply');
                    });
                    if (applyBtn) { applyBtn.click(); }
                }
            });
        }
        if (document.readyState === 'loading') {
            document.addEventListener('DOMContentLoaded', hookCtrlEnter);
        } else {
            hookCtrlEnter();
        }
    })();
    </script>
    """, unsafe_allow_html=True)

    col_title, col_reset = st.columns([5, 1])
    with col_title:
        st.markdown(f"## 👤 تحليل الموظف: {name} (ID: {eid})")
    with col_reset:
        st.markdown("<div style='padding-top:18px'>", unsafe_allow_html=True)
        if st.button("🔄 استعادة الافتراضي", key=f"reset_emp_{eid}",
                ):
            st.session_state[f"confirm_reset_{eid}"] = True
        st.markdown("</div>", unsafe_allow_html=True)

    # ── حوار تأكيد الاستعادة ───────────────────────────────────────────
    if st.session_state.get(f"confirm_reset_{eid}"):
        # نحسب كم تصحيح سيُحذف
        n_corr = len(st.session_state.get('corrections', {}).get(eid, {}))
        n_corr += sum(
            1 for k in st.session_state.get('corrections', {})
            if k.startswith(f"{eid}_")
        )
        n_ov   = sum(1 for k in st.session_state.get('day_overrides', {}) if k.startswith(f"{eid}_"))
        n_pend = sum(
            1 for k in st.session_state.get('pending_changes', {})
            if k.startswith(f"corr_{eid}_") or k.startswith(f"dayov_{eid}_")
        )

        st.warning(
            f"⚠️ سيتم حذف **كل** تصحيحات وتعديلات الموظف **{name}** "
            f"({n_ov} تعديل يوم، {n_pend} معلّق). هذا لا يمكن التراجع عنه!"
        )
        rc1, rc2, rc3 = st.columns([2, 2, 4])
        with rc1:
            if st.button("✅ نعم، استعادة الافتراضي", key=f"reset_confirm_{eid}", type="primary"):
                # حذف corrections الخاصة بهذا الموظف
                corr = st.session_state.get('corrections', {})
                keys_to_del = [k for k in corr if k.startswith(f"{eid}_")]
                for k in keys_to_del:
                    del corr[k]
                st.session_state.corrections = corr
                # حذف day_overrides
                ov_dict = st.session_state.get('day_overrides', {})
                ov_keys = [k for k in ov_dict if k.startswith(f"{eid}_")]
                for k in ov_keys:
                    del ov_dict[k]
                st.session_state.day_overrides = ov_dict
                # حذف pending changes
                pend = st.session_state.get('pending_changes', {})
                pend_keys = [
                    k for k in pend
                    if k.startswith(f"corr_{eid}_") or k.startswith(f"dayov_{eid}_")
                       or k.startswith(f"dayov_remove_{eid}_")
                ]
                for k in pend_keys:
                    del pend[k]
                st.session_state.pending_changes = pend
                st.session_state[f"confirm_reset_{eid}"] = False
                st.success(f"✅ تمت استعادة الإعدادات الافتراضية للموظف {name}")
                st.rerun()
        with rc2:
            if st.button("❌ إلغاء", key=f"reset_cancel_{eid}"):
                st.session_state[f"confirm_reset_{eid}"] = False
                st.rerun()

    ov          = overrides_summary.get(eid, {})
    work_h      = ov.get('work_hours',      emp_row.get('work_hours', 0))
    att_days    = ov.get('attendance_days', emp_row.get('attendance_days', 0))
    absent_days = ov.get('absent_days',     emp_row.get('absent_days', 0))
    inc_days    = ov.get('incomplete_days', emp_row.get('incomplete_days', 0))
    avg_min     = round(work_h * 60 / att_days) if att_days > 0 else 0

    col1, col2, col3, col4, col5 = st.columns(5)
    with col1: st.metric("أيام الحضور",        fmt_days(att_days))
    with col2: st.metric("أيام الغياب",        int(absent_days))
    with col3: st.metric("ساعات العمل الكلية",  fmt_h(work_h))
    with col4: st.metric("أيام بصمة ناقصة",    int(inc_days))
    with col5: st.metric("متوسط ساعات/يوم",    fmt_hm(avg_min))

    if not days:
        st.warning("لا توجد بيانات يومية لهذا الموظف")
        return

    # نحسب عدد البصمات الناقصة الفعلية (غير محلولة)
    corrections = st.session_state.get('corrections', {})
    pending     = st.session_state.get('pending_changes', {})

    unresolved_count = 0
    for di, d in enumerate(days):
        if not d.get('needs_review'):
            continue
        key_base = f"{eid}_{di}"
        if key_base in st.session_state.get('day_overrides', {}):
            continue
        for ri, rev in enumerate(d['needs_review']):
            rev_key = f"{key_base}_r{ri}"
            if corrections.get(key_base, {}).get(rev_key):
                continue
            pending_key = f"corr_{key_base}_{ri}"
            if pending_key in pending:
                continue
            unresolved_count += 1

    if unresolved_count > 0:
        with st.expander(f"🔧 مراجعة البصمات الناقصة ({unresolved_count} بصمة تحتاج مراجعة)", expanded=True):
            show_review_panel(eid, days)
    else:
        st.success("✅ لا توجد بصمات ناقصة — جميع البصمات مكتملة أو تم تصحيحها")

    # ── محرر يوم مباشر من الجدول ──
    quick_edit_key = f"quick_edit_day_{eid}"
    if quick_edit_key in st.session_state and st.session_state[quick_edit_key] is not None:
        qdi = st.session_state[quick_edit_key]
        qd  = days[qdi]
        st.markdown(f"### ✏️ تعديل يوم {qd['day']} مباشرة")
        col_back, _ = st.columns([1, 5])
        with col_back:
            if st.button("← إغلاق التعديل", key=f"close_quick_{eid}"):
                # تحقق فعلي من وجود تعديلات معلّقة قبل الإغلاق
                if count_pending_changes() > 0:
                    st.session_state['leaving_to_main'] = True
                else:
                    st.session_state[quick_edit_key] = None
                st.rerun()
        show_day_editor(eid, days, preselect_di=qdi)
        st.markdown("---")
    else:
        with st.expander("🛠️ تعديل يوم بالكامل (تصحيح شامل لأي يوم)", expanded=False):
            show_day_editor(eid, days)

    # ── عرض الجدول ──
    day_overrides = st.session_state.get('day_overrides', {})

    # دمج pending day overrides في العرض
    pending_overrides_display = {}
    for pk, pv in pending.items():
        if pv.get('type') == 'day_override' and pk.startswith(f"dayov_{eid}_"):
            kb = pv['key_base']
            pending_overrides_display[kb] = pv['data']

    result = []
    for di, d in enumerate(days):
        key_base = f"{eid}_{di}"

        # هل هذا اليوم لديه override مؤكد أو معلّق؟
        ov_day = day_overrides.get(key_base) or pending_overrides_display.get(key_base)
        is_pending_ov = key_base in pending_overrides_display and key_base not in day_overrides

        if ov_day:
            status    = ov_day.get('status')
            pairs     = ov_day.get('pairs') or []
            day_total = 0
            pairs_disp = []
            for ci, co in pairs:
                diff = minutes_diff(ci, co)
                if diff is not None:
                    day_total += diff
                    pairs_disp.append((f"{ci}→{co}", diff))

            punches = ("  |  ".join(f"{fmt_pair_12h(p[0])} ({fmt_hm(p[1])})" for p in pairs_disp)
                       if pairs_disp else "—")

            note = "⏳ معلّق" if is_pending_ov else "✏️ معدَّل يدوياً"
            result.append({
                'اليوم':          d['day'],
                'الحالة':         status,
                'فترات الحضور':   punches,
                'إجمالي الساعات': fmt_hm(day_total) if day_total > 0 else "—",
                'ملاحظة':         note,
                '_min':           day_total,
                '_status':        status,
                '_incomplete':    False,
                '_corrected':     False,
                '_overridden':    True,
                '_pending':       is_pending_ov,
            })
            continue

        # حساب تلقائي
        day_corrections = corrections.get(key_base, {})

        base_min      = sum(m for _, m in d['punch_pairs'])
        extra_min     = 0
        # كل عنصر: (pair_str, diff, corrected_side) — corrected_side يحدد أي طرف
        # من الزوج تم إدخاله/اقتراحه يدويًا (وليس بصمة فعلية من الجهاز)
        pairs_display = [(p[0], p[1], None) for p in d['punch_pairs']]

        for ri, rev in enumerate(d.get('needs_review', [])):
            rev_key       = f"{key_base}_r{ri}"
            resolved_time = day_corrections.get(rev_key)

            # هل هناك تصحيح معلّق؟
            pending_corr_key = f"corr_{key_base}_{ri}"
            pending_corr = pending.get(pending_corr_key, {}).get('value') if pending_corr_key in pending else None

            actual_resolved = resolved_time or pending_corr
            if actual_resolved:
                # تحديد punch_role: هل البصمة المعروفة (ci) حضور أم انصراف؟
                punch_role = day_corrections.get(f"{rev_key}_role")
                if punch_role is None and pending_corr_key in pending:
                    punch_role = pending.get(pending_corr_key, {}).get('punch_role')
                if punch_role is None:
                    punch_role = 'check_in'

                if punch_role == 'check_out':
                    # ci = انصراف (بصمة حقيقية)، actual_resolved = حضور (مُصحَّح)
                    diff = minutes_diff(actual_resolved, rev['ci'])
                    pair_str = f"{actual_resolved}→{rev['ci']}"
                    corrected_side = 'first'
                else:
                    # ci = حضور (بصمة حقيقية)، actual_resolved = انصراف (مُصحَّح، الافتراضي)
                    diff = minutes_diff(rev['ci'], actual_resolved)
                    pair_str = f"{rev['ci']}→{actual_resolved}"
                    corrected_side = 'second'

                if diff is not None:
                    extra_min += diff
                    mark = "⏳✓" if pending_corr and not resolved_time else "✓"
                    pairs_display.append((f"{pair_str}{mark}", diff, corrected_side))

        day_total        = base_min + extra_min
        still_incomplete = d['incomplete'] and any(
            not day_corrections.get(f"{key_base}_r{ri}") and
            f"corr_{key_base}_{ri}" not in pending
            for ri in range(len(d.get('needs_review', [])))
        )

        # ── عرض البصمة الوحيدة الموجودة وتحديد مكان البصمة الناقصة ──
        missing_punch_display = []
        if still_incomplete:
            _saved_fp_types = st.session_state.get('fingerprint_types', {}).get(eid, {})
            for ri, rev in enumerate(d.get('needs_review', [])):
                rev_key = f"{key_base}_r{ri}"
                if day_corrections.get(rev_key):
                    continue
                pending_corr_key = f"corr_{key_base}_{ri}"
                if pending_corr_key in pending:
                    continue

                ci = rev.get('ci', '')
                if not ci:
                    continue

                saved_type = _saved_fp_types.get(rev_key)
                if saved_type:
                    p_type = saved_type
                else:
                    p_type = classify_lone_punch(
                        ci, days, shift_pattern=d.get('shift_pattern')
                    )['type']

                ci_disp = to_12h(ci)
                if p_type == 'check_out':
                    missing_punch_display.append(f"\u202b⚠️ ؟ ← {ci_disp}\u202c")
                else:
                    missing_punch_display.append(f"\u202b⚠️ {ci_disp} ← ؟\u202c")

        if pairs_display:
            punches = "  |  ".join(
                f"{fmt_pair_12h_html(p[0], p[2])} ({fmt_hm(p[1])})" for p in pairs_display
            )
            if missing_punch_display:
                punches += "  |  " + "  |  ".join(missing_punch_display)
        elif missing_punch_display:
            punches = "  |  ".join(missing_punch_display)
        else:
            punches = "—"

        # إظهار التسامح المضاف
        tol_added = d.get('tolerance_added', 0)
        if tol_added > 0:
            note = f"⚖️ +{tol_added} د"
        else:
            note = ""

        if still_incomplete:
            note = "⚠️ بصمة ناقصة" + (" " + note if note else "")
        elif d['incomplete']:
            note = "✅ تم التصحيح" + (" " + note if note else "")

        # ── تحديد الحالة الفعلية لليوم بعد التصحيح ───────────────────────
        # نفس منطق apply_overrides بالضبط: لو اليوم كان "غياب" في الملف
        # الخام لكن أصبح له ساعات عمل فعلية بعد تصحيح البصمة الناقصة،
        # يُعرض "حضور" - وليس "غياب" كما كان يحدث سابقًا (لأن العرض كان
        # يعتمد دائمًا على d['status'] الأصلي بدون النظر لنتيجة التصحيح).
        effective_status = d['status']
        if effective_status == 'غياب' and day_total > 0:
            effective_status = 'حضور'

        result.append({
            'اليوم':          d['day'],
            'الحالة':         effective_status,
            'فترات الحضور':   punches,
            'إجمالي الساعات': fmt_hm(day_total) if day_total > 0 else "—",
            'ملاحظة':         note,
            '_min':           day_total,
            '_status':        effective_status,
            '_incomplete':    still_incomplete,
            '_corrected':     d['incomplete'] and not still_incomplete,
            '_overridden':    False,
            '_pending':       False,
        })

    daily_df = pd.DataFrame(result)

    present = daily_df[daily_df['_min'] > 0]
    if not present.empty:
        st.markdown("### 📊 ساعات العمل اليومية")
        st.bar_chart((present.set_index('اليوم')['_min'] / 60).rename("ساعات العمل"))

    st.markdown("### 📅 تفاصيل الحضور اليومي")

    quick_edit_key = f"quick_edit_day_{eid}"
    if quick_edit_key not in st.session_state:
        st.session_state[quick_edit_key] = None

    display_cols = ['اليوم', 'الحالة', 'فترات الحضور', 'إجمالي الساعات', 'ملاحظة']

    def _row_colors(row):
        status     = row['_status']
        incomp     = row['_incomplete']
        corrected  = row['_corrected']
        overridden = row['_overridden']
        is_pending = row['_pending']
        if is_pending:
            return '#faf1de', '#8a6a10'   # amber cream للمعلّق
        elif overridden:
            return '#eef2ee', '#2f6f63'   # teal-tinted cream للمُعدَّل
        elif status == 'غياب':
            return '#fbeceb', '#c64545'   # red-tinted cream
        elif incomp:
            return '#faf1de', '#8a6a10'   # amber cream
        elif corrected:
            return '#eaf5ec', '#3c7a4c'   # green-tinted cream
        else:
            return '#faf9f5', '#252523'   # default canvas

    # ── نبني الجدول كـ HTML صريح (بدل st.dataframe) عشان نقدر نعرض البصمة
    # المُصحَّحة Bold وبلون مختلف داخل نفس الخلية — الجداول العادية في Streamlit
    # (glide-data-grid) بترسم النص كـ canvas ومش بتدعم HTML/Bold جوه الخلية.
    import html as _html_mod

    _header_html = "".join(
        f"<th style='padding:0.65rem 0.85rem;text-align:right;color:#6c6a64;"
        f"font-size:0.78rem;font-weight:600;text-transform:uppercase;"
        f"letter-spacing:0.05em;border-bottom:1px solid #e6dfd8;background:#efe9de;'>"
        f"{_html_mod.escape(c)}</th>"
        for c in display_cols
    )

    _rows_html = []
    for _, _row in daily_df.iterrows():
        _bg, _fg = _row_colors(_row)
        _cells = []
        for _c in display_cols:
            _val = _row[_c]
            # 'فترات الحضور' مبني مسبقًا كـ HTML (فيه Bold للبصمة المُصحَّحة) —
            # باقي الأعمدة نصوص عادية بنعمل لها escape للأمان
            _cell_content = _val if _c == 'فترات الحضور' else _html_mod.escape(str(_val))
            _cells.append(
                f"<td style='padding:0.55rem 0.85rem;"
                f"border-bottom:1px solid rgba(230,223,216,0.7);"
                f"text-align:right;color:{_fg};white-space:nowrap;'>{_cell_content}</td>"
            )
        _rows_html.append(f"<tr style='background-color:{_bg};'>{''.join(_cells)}</tr>")

    _table_html = (
        "<div style='overflow-x:auto;border:1px solid #e6dfd8;border-radius:12px;'>"
        "<table style='width:100%;border-collapse:collapse;direction:rtl;font-size:0.84rem;'>"
        f"<thead><tr>{_header_html}</tr></thead>"
        f"<tbody>{''.join(_rows_html)}</tbody>"
        "</table></div>"
    )
    st.markdown(_table_html, unsafe_allow_html=True)


    all_rows_data = list(daily_df.iterrows())
    ROW_SIZE = 10
    for row_start in range(0, len(all_rows_data), ROW_SIZE):
        chunk = all_rows_data[row_start: row_start + ROW_SIZE]
        btn_cols = st.columns(len(chunk))
        for ci, (row_idx, row) in enumerate(chunk):
            day_num_str = row['اليوم']
            di_for_row = next(
                (di for di, d in enumerate(days) if d['day'] == day_num_str),
                None
            )
            if di_for_row is not None:
                is_active  = (st.session_state[quick_edit_key] == di_for_row)
                overridden = daily_df.at[row_idx, '_overridden']
                incomp     = daily_df.at[row_idx, '_incomplete']
                is_pend    = daily_df.at[row_idx, '_pending']
                icon = "🔒" if is_active else ("⏳✏️" if is_pend else ("✅✏️" if overridden else ("⚠️✏️" if incomp else "✏️")))
                with btn_cols[ci]:
                    st.caption(f"**{day_num_str}**")
                    if st.button(icon, key=f"qedit_{eid}_{di_for_row}"):
                        st.session_state[quick_edit_key] = None if is_active else di_for_row
                        st.rerun()

    # ── شريط Apply للموظف الحالي ──
    show_float_apply_bar()


# ─── Bulk Smart Apply (all employees) ────────────────────────────────────────

def has_chronological_conflict(d: dict, punch_type: str, corr_suggestion) -> bool:
    """
    فحص "زيرو تسامح": يتأكد إن الاقتراح الذكي لا يكسر الترتيب الزمني الفعلي
    لبصمات نفس اليوم قبل ما يُطبَّق تلقائيًا في bulk_smart_apply_all.

    البصمة الناقصة (لسه من غير زوج) هي دائمًا آخر عنصر في raw_times.
    بنقارن:
      1) البصمة الناقصة نفسها لازم تكون بعد الجار السابق ليها (آخر بصمة قبلها).
      2) الاقتراح المحسوب لازم يحافظ على نفس الترتيب:
         - لو النوع check_in  → الاقتراح (انصراف) لازم يكون بعد البصمة الناقصة.
         - لو النوع check_out → الاقتراح (حضور) لازم يكون بعد الجار السابق وقبل البصمة الناقصة.

    ترجع True لو فيه تعارض (يعني نرفض الاقتراح)، وFalse لو كل حاجة متسقة.
    """
    raw_times = d.get('raw_times', [])
    if len(raw_times) < 1:
        return True  # مفيش بيانات كافية أصلاً

    lone_min = to_minutes(raw_times[-1])
    if lone_min is None:
        return True

    prev_min = to_minutes(raw_times[-2]) if len(raw_times) >= 2 else None

    # 1) البصمة الناقصة نفسها لازم تيجي بعد الجار السابق مباشرة
    if prev_min is not None and lone_min <= prev_min:
        return True

    if not corr_suggestion:
        return False  # مفيش اقتراح أصلاً هيتفحص - القرار في مكان تاني

    corr_min = to_minutes(corr_suggestion)
    if corr_min is None:
        return True

    if punch_type == 'check_in':
        # الاقتراح = وقت انصراف، لازم بعد البصمة الناقصة (بسماح عبور منتصف الليل)
        diff = corr_min - lone_min
        if diff < 0:
            diff += 1440
        if not (0 < diff <= 960):
            return True

    elif punch_type == 'check_out':
        # الاقتراح = وقت حضور، لازم بعد الجار السابق وقبل البصمة الناقصة
        if prev_min is not None and not (prev_min < corr_min < lone_min):
            return True
        if prev_min is None and not (corr_min < lone_min):
            return True

    else:
        return True  # نوع غير معروف - رفض احتياطي

    return False


def bulk_smart_apply_all(emp_days, min_sample: int = 3):
    """
    يجمع كل البصمات الناقصة لكل الموظفين ويطبّق الاقتراح الذكي فقط إذا:
      - confidence == 'high'
      - corr_sample_size >= min_sample
      - لم تُحلَّ بعد (لا في corrections ولا في pending ولا في day_overrides)

    يُرجع dict يصف ما تم وما تُرك:
      'applied':  list of dicts  (ما طُبِّق)
      'skipped':  list of dicts  (ما تُرك ولماذا)
    """
    corrections   = st.session_state.get('corrections', {})
    pending       = st.session_state.get('pending_changes', {})
    day_overrides = st.session_state.get('day_overrides', {})

    applied = []
    skipped = []

    for eid, days in emp_days.items():
        for di, d in enumerate(days):
            if not d.get('needs_review'):
                continue
            key_base = f"{eid}_{di}"
            if key_base in day_overrides:
                continue
            for ri, rev in enumerate(d['needs_review']):
                rev_key     = f"{key_base}_r{ri}"
                pending_key = f"corr_{key_base}_{ri}"

                # تخطّي المحلولة مسبقاً
                if corrections.get(key_base, {}).get(rev_key):
                    continue
                if pending_key in pending:
                    continue

                ci  = rev.get('ci', '')
                sp  = d.get('shift_pattern')
                clf = classify_lone_punch(ci, days, shift_pattern=sp)

                punch_type       = clf['type']
                confidence       = clf['confidence']
                corr_suggestion  = clf['corr_suggestion']
                corr_sample_size = clf['corr_sample_size']

                info_base = {
                    'eid':      eid,
                    'day':      d['day'],
                    'ci':       ci,
                    'type':     punch_type,
                    'conf':     confidence,
                    'sample':   corr_sample_size,
                    'suggest':  corr_suggestion,
                }

                # شروط التطبيق الآمن
                if punch_type not in ('check_in', 'check_out'):
                    skipped.append({**info_base, 'reason': 'نوع غير معروف'})
                    continue
                if confidence != 'high':
                    skipped.append({**info_base, 'reason': 'ثقة منخفضة 🟡'})
                    continue
                if not corr_suggestion:
                    skipped.append({**info_base, 'reason': 'لا يوجد اقتراح'})
                    continue
                if corr_sample_size < min_sample:
                    skipped.append({**info_base, 'reason': f'عينة صغيرة ({corr_sample_size} أيام)'})
                    continue
                if has_chronological_conflict(d, punch_type, corr_suggestion):
                    skipped.append({**info_base, 'reason': '⛔ تعارض في الترتيب الزمني'})
                    continue

                # كل الشروط مرّت — نضيفه للـ pending
                add_pending_change(pending_key, {
                    'type':       'correction',
                    'key_base':   key_base,
                    'rev_key':    rev_key,
                    'value':      corr_suggestion,
                    'punch_role': punch_type,
                })
                applied.append(info_base)

    return {'applied': applied, 'skipped': skipped}


# ─── Main UI ──────────────────────────────────────────────────────────────────

_top_col1, _top_col2 = st.columns([5, 1])
with _top_col1:
    st.title("💼 نظام إدارة الحضور والرواتب")
    if st.session_state.get('_is_anonymous'):
        st.caption("📤 مراجعة مؤقتة (غير محفوظة) — نزّل تقريرك أولاً ثم أغلق البرنامج")
        if st.button("🚪 إنهاء المراجعة وإغلاق البرنامج", key="_anon_close_btn", type="secondary"):
            st.info("👋 جاري إغلاق البرنامج...")
            import os as _os_close
            _os_close._exit(0)
    elif st.session_state.get('_current_company') and st.session_state.get('_current_year') and st.session_state.get('_current_month'):
        _co_t  = st.session_state['_current_company']
        _mo_t  = _ARABIC_MONTHS[st.session_state['_current_month']]
        _yr_t  = st.session_state['_current_year']
        st.caption(f"📂 {_co_t} — {_mo_t} {_yr_t}")
with _top_col2:
    st.write("")
    with st.popover(f"👤 {st.session_state['license_client'].get('client_name', 'مستخدم')}", use_container_width=True):
        st.markdown(f"**اسم المستخدم:** {st.session_state['license_client'].get('client_name', '')}")

# ── حفظ واسترجاع موضع الـ Scroll بعد كل rerun ──────────────────────────
st.markdown("""
<script>
(function() {
    // انتظر حتى يكتمل تحميل Streamlit
    function preserveScroll() {
        const SCROLL_KEY = 'st_scroll_pos';
        const mainEl = window.parent.document.querySelector('.main') ||
                       window.parent.document.querySelector('[data-testid="stMain"]') ||
                       window.parent.document.documentElement;

        // استرجع الموضع المحفوظ فوراً قبل أي شيء
        const saved = sessionStorage.getItem(SCROLL_KEY);
        if (saved !== null) {
            const pos = parseInt(saved, 10);
            requestAnimationFrame(() => {
                mainEl.scrollTop = pos;
                window.parent.scrollTo(0, pos);
                // محاولة ثانية بعد render
                setTimeout(() => {
                    mainEl.scrollTop = pos;
                    window.parent.scrollTo(0, pos);
                }, 150);
            });
        }

        // احفظ الموضع عند كل scroll
        const trackEl = window.parent.document;
        if (!trackEl._scrollTracked) {
            trackEl._scrollTracked = true;
            window.parent.addEventListener('scroll', function() {
                sessionStorage.setItem(SCROLL_KEY, window.parent.scrollY);
            }, { passive: true });
            // تتبع الـ main container برضو
            if (mainEl !== window.parent.document.documentElement) {
                mainEl.addEventListener('scroll', function() {
                    sessionStorage.setItem(SCROLL_KEY, mainEl.scrollTop);
                }, { passive: true });
            }
        }
    }

    if (document.readyState === 'loading') {
        document.addEventListener('DOMContentLoaded', preserveScroll);
    } else {
        preserveScroll();
    }
})();
</script>
""", unsafe_allow_html=True)
st.markdown("---")

# ── تهيئة session_state للجلسة الحالية (قبل السايدبار عشان السلايدرز تقدر تقرأها) ──
if '_current_company' not in st.session_state:
    st.session_state['_current_company'] = None
if '_current_year' not in st.session_state:
    st.session_state['_current_year'] = None
if '_current_month' not in st.session_state:
    st.session_state['_current_month'] = None
if '_is_anonymous' not in st.session_state:
    st.session_state['_is_anonymous'] = False
if '_show_import_popup' not in st.session_state:
    st.session_state['_show_import_popup'] = False
if '_confirm_reset' not in st.session_state:
    st.session_state['_confirm_reset'] = False
if '_tolerance_enabled' not in st.session_state:
    st.session_state['_tolerance_enabled'] = False
if '_tolerance_minutes' not in st.session_state:
    st.session_state['_tolerance_minutes'] = 20

# ── الإعدادات المحفوظة الخاصة بالملف الحالي (لو موجود) — تُستخدم كـ default للسلايدرز ──
_co_now = st.session_state.get('_current_company')
_yr_now = st.session_state.get('_current_year')
_mo_now = st.session_state.get('_current_month')
if _co_now and _yr_now and _mo_now:
    _saved_settings = load_month_settings(_co_now, _yr_now, _mo_now)
else:
    _saved_settings = {}
_default_cutoff   = _saved_settings.get('redistribute_minutes', 180)
_default_sat_on   = _saved_settings.get('saturate_enabled', False)
_default_sat_min  = _saved_settings.get('saturate_minutes', 60)
_default_device_idx = _saved_settings.get('device_type_idx', 0)
_default_tol_on   = _saved_settings.get('tolerance_enabled', False)
_default_tol_min  = _saved_settings.get('tolerance_minutes', 20)
_default_dup_min  = _saved_settings.get('duplicate_punch_tolerance', 10)

with st.sidebar:
    tab_settings, tab_files = st.tabs(["⚙️ الإعدادات", "📁 ملفاتي"])

with tab_settings:
    # مفتاح فريد لكل ملف (شركة/سنة/شهر) عشان أي widget يرجع لقيمته المحفوظة
    # الخاصة بيه بالضبط عند التبديل بين الملفات، مش يفضل محتفظ بقيمة ملف سابق
    _file_key_suffix = f"{_co_now}_{_yr_now}_{_mo_now}" if _co_now else "none"

    st.markdown("## 🖥️ نوع جهاز البصمة")
    _device_radio_key = f"device_type_radio_{_file_key_suffix}"
    if _device_radio_key not in st.session_state:
        st.session_state[_device_radio_key] = ["Main Fingerprint", "HikVision"][_default_device_idx]
    device_type = st.radio(
        "اختر نوع الجهاز",
        options=["Main Fingerprint", "HikVision"],
        key=_device_radio_key
    )
    st.session_state['_device_type_idx'] = 0 if device_type == "Main Fingerprint" else 1

    # ── حفظ نوع الجهاز فورًا كإعداد خاص بهذا الملف بالذات ──
    if _co_now and not st.session_state.get('_is_anonymous'):
        save_month_settings(_co_now, _yr_now, _mo_now, {
            **_saved_settings,
            'device_type_idx': st.session_state['_device_type_idx'],
        })

    

    st.markdown("---")
    st.markdown("## ⚙️ إعدادات الشيفتات")

    # ── حد ما بعد منتصف الليل (بالدقائق، خطوة 10) ──
    # نستخدم select_slider بدل slider العادي عشان نعرض القيمة بصيغة
    # "ساعة:دقيقة" (مثل 3:00) بدل رقم خام صعب القراءة (مثل 180)

    _cutoff_options = list(range(0, OVERNIGHT_WINDOW_MAX + 1, 10))
    _cutoff_key = f"cutoff_slider_{_file_key_suffix}"
    if _cutoff_key not in st.session_state:
        st.session_state[_cutoff_key] = _default_cutoff if _default_cutoff in _cutoff_options else min(_cutoff_options, key=lambda x: abs(x - _default_cutoff))
    cutoff_raw = st.select_slider(
        "🌙 حد بصمات ما بعد منتصف الليل",
        options=_cutoff_options,
        key=_cutoff_key,
        format_func=lambda m: f"{m // 60}:{m % 60:02d}"
    )
    cutoff_hour = cutoff_raw / 60

    st.markdown("")

    # ── حد تشبع نهاية اليوم (بالدقائق، خطوة 10) ──
    _sat_enabled_key = f"sat_enabled_{_file_key_suffix}"
    if _sat_enabled_key not in st.session_state:
        st.session_state[_sat_enabled_key] = _default_sat_on
    saturate_enabled = st.checkbox(
        "⏰ تفعيل حد نهاية اليوم (Saturate)",
        key=_sat_enabled_key
    )
    if saturate_enabled:
        _saturate_options = list(range(0, OVERNIGHT_WINDOW_MAX + 1, 10))
        _sat_key = f"sat_slider_{_file_key_suffix}"
        if _sat_key not in st.session_state:
            st.session_state[_sat_key] = _default_sat_min if _default_sat_min in _saturate_options else min(_saturate_options, key=lambda x: abs(x - _default_sat_min))
        saturate_raw = st.select_slider(
            "⏰ حد نهاية اليوم (Saturate)",
            options=_saturate_options,
            key=_sat_key,
            format_func=lambda m: f"{m // 60}:{m % 60:02d}",
            help=(
                "دالة مستقلة تمامًا عن حد منتصف الليل أعلاه — تعمل فقط على "
                "البصمات الواقعة بين 12:00 ص و6:00 ص.\n\n"
                "أي بصمة في هذه الفترة تتجاوز هذا الحد تُستبدل بهذا الحد نفسه. "
                "بصمات النهار/المساء الطبيعية لا تتأثر أبدًا.\n\n"
                "مثال: حد منتصف الليل = 4:30 ص، وهذا الحد = 1:00 ص → "
                "بصمة 3:45 ص تُحسب كـ 1:00 ص، أما بصمة 8:00 ص فتبقى كما هي."
            )
        )
        sh, sm = divmod(saturate_raw, 60)
        wh, wm = divmod(OVERNIGHT_WINDOW_MAX, 60)
        st.caption(
            f"بصمات بين 12:00 ص و{int(wh)}:{wm:02d} ص بعد {int(sh)}:{sm:02d} ص "
            f"→ تُستبدل بـ {int(sh)}:{sm:02d} ص"
        )
    else:
        saturate_raw = None
        st.caption("معطَّل — البصمات تُحسب كما هي بعد Redistribute")

    # ═══════════ إعدادات التسامح مع المغادرة المبكرة ═══════════
    st.markdown("---")
    st.markdown("## ⚖️ تعويض المغادرة المبكرة")
    _tol_enabled_key = f"tol_enabled_{_file_key_suffix}"
    if _tol_enabled_key not in st.session_state:
        st.session_state[_tol_enabled_key] = _default_tol_on
    tolerance_enabled = st.checkbox(
        "تفعيل تعويض المغادرة المبكرة",
        key=_tol_enabled_key
    )
    if tolerance_enabled:
        _tol_options = list(range(0, 61, 5))
        _tol_key = f"tol_slider_{_file_key_suffix}"
        if _tol_key not in st.session_state:
            st.session_state[_tol_key] = _default_tol_min if _default_tol_min in _tol_options else min(_tol_options, key=lambda x: abs(x - _default_tol_min))
        tolerance_minutes = st.select_slider(
            "أقصى مدة تسامح (دقيقة)",
            options=_tol_options,
            key=_tol_key,
            format_func=lambda m: f"{m} دقيقة"
        )
    else:
        tolerance_minutes = 0

    # ═══════════ إعدادات البصمات المتقاربة ═══════════
    st.markdown("---")
    st.markdown("## 🔄 حذف البصمات المتقاربة")

    _dup_key = f"dup_punch_slider_{_file_key_suffix}"
    if _dup_key not in st.session_state:
        st.session_state[_dup_key] = _default_dup_min

    duplicate_punch_tolerance = st.slider(
        "⏱️ الفارق المسموح بين البصمات المتتالية (دقائق)",
        min_value=1,
        max_value=60,
        value=st.session_state[_dup_key],
        step=1,
        key=_dup_key,
        help=(
            "إذا بصم الموظف مرتين بفاصل أقل من هذا الحد، يتم حذف البصمة الثانية.\n\n"
            "مثال: إذا كان الحد 10 دقائق وبصم الموظف في 8:00 و8:05، "
            "سيتم حذف البصمة 8:05 ويبقى فقط 8:00"
        )
    )
    st.caption(f"الحد الحالي: {duplicate_punch_tolerance} دقيقة")

    # ── حفظ الإعدادات فورًا كإعدادات خاصة بهذا الملف بالذات ──
    if _co_now and not st.session_state.get('_is_anonymous'):
        save_month_settings(_co_now, _yr_now, _mo_now, {
            'redistribute_minutes': cutoff_raw,
            'saturate_enabled': saturate_enabled,
            'saturate_minutes': saturate_raw if saturate_raw is not None else _default_sat_min,
            'device_type_idx': st.session_state.get('_device_type_idx', _default_device_idx),
            'tolerance_enabled': tolerance_enabled,
            'tolerance_minutes': tolerance_minutes,
            'duplicate_punch_tolerance': duplicate_punch_tolerance,
        })
    # تخزين القيم في session_state لاستخدامها في apply_overrides
    st.session_state['_tolerance_enabled'] = tolerance_enabled
    st.session_state['_tolerance_minutes'] = tolerance_minutes
    st.session_state['_saturate_minutes'] = saturate_raw
    st.session_state['_duplicate_punch_tolerance'] = duplicate_punch_tolerance

    st.markdown("---")

    # عرض عدد التعديلات المعلّقة في الـ Sidebar أيضاً
    n_pending = count_pending_changes()
    if n_pending > 0:
        st.warning(f"⏳ {n_pending} تعديل معلّق بانتظار Apply")
        if st.button("✅ Apply all", key="sidebar_apply", type="primary"):
            count = apply_all_pending()
            st.success(f"✅ تم تطبيق {count} تعديل!")
            st.rerun()
        if st.button("🗑️ تجاهل الكل", key="sidebar_discard"):
            discard_all_pending()
            st.rerun()

def _open_month_file(company: str, year: int, month: int):
    """فتح ملف شهر معيّن وتحميل كل إعداداته (نفس منطق زر 'فتح آخر ملف')."""
    _bytes = load_work_file_bytes(company, year, month)
    if not _bytes:
        st.error("❌ لم يتم العثور على الملف")
        return
    st.session_state['_cached_file_bytes'] = _bytes
    st.session_state['_current_company']  = company
    st.session_state['_current_year']     = year
    st.session_state['_current_month']    = month
    st.session_state['_is_anonymous']     = False
    st.session_state['corrections']       = {}
    st.session_state['day_overrides']     = {}
    st.session_state['restore_done']      = False
    st.session_state['file_state_loaded'] = False
    st.session_state['last_file_hash']    = None
    _defs = load_employee_defaults(company)
    st.session_state['saved_rates'] = _defs if _defs else {}
    (Path(_get_data_root()) / "last_company.txt").write_text(company, encoding="utf-8")
    st.session_state['_show_import_popup'] = False
    _full_rerun()

with tab_files:
    # ── ستايل قوي يشيل خلفية/حدود كل الأزرار في السايدبار عشان تبان
    # كصفوف شجرة مضغوطة بدون كروت أو خلفيات (زي Notion) ──
    st.markdown("""
        <style>
        section[data-testid="stSidebar"] button:not([kind="primary"]) {
            background: transparent !important;
            border: none !important;
            box-shadow: none !important;
            padding: 0.1rem 0.35rem !important;
            min-height: 1.7rem !important;
            font-size: 0.85rem !important;
            text-align: left !important;
            justify-content: flex-start !important;
            border-radius: 4px !important;
        }
        section[data-testid="stSidebar"] button:not([kind="primary"]):hover {
            background: rgba(255,255,255,0.08) !important;
        }
        section[data-testid="stSidebar"] div[data-testid="stPopover"] button {
            justify-content: center !important;
            padding: 0.1rem 0.3rem !important;
            width: auto !important;
        }
        section[data-testid="stSidebar"] div[data-testid="column"] {
            gap: 0.1rem !important;
        }
        </style>
    """, unsafe_allow_html=True)

    _IND1 = "\u00A0" * 4   # مسافة بادئة مستوى السنة
    _IND2 = "\u00A0" * 8   # مسافة بادئة مستوى الشهر

    @_st_fragment
    def _render_file_tree():
        _tree = list_company_tree()
        if not _tree:
            st.caption("لا توجد ملفات محفوظة بعد.")
            return

        for _company in sorted(_tree.keys()):
            _co_open_key = f"_tree_co_open_{_company}"
            if _co_open_key not in st.session_state:
                st.session_state[_co_open_key] = (_company == _co_now)

            _co_arrow = ""
            _co_current_mark = " 🔵" if _company == _co_now else ""

            # ── صف الشركة: تبديل الفتح/الطي (سريع بدون Rerun لكامل الصفحة) + قائمة ⋮ ──
            _co_lbl, _co_menu = st.columns([6, 1], gap="small")
            with _co_lbl:
                if st.button(
                    f"{_co_arrow} 🏢 {_company}{_co_current_mark}",
                    key=f"_co_toggle_{_company}", use_container_width=True
                ):
                    st.session_state[_co_open_key] = not st.session_state[_co_open_key]
                    st.rerun()
            with _co_menu:
                with st.popover("⋮", use_container_width=True):
                    if st.session_state.get('_renaming_company') == _company:
                        _new_name = st.text_input(
                            "الاسم الجديد", value=_company, key=f"_rename_input_{_company}"
                        )
                        _rn_a, _rn_b = st.columns(2)
                        with _rn_a:
                            if st.button("✅ حفظ", key=f"_rename_save_{_company}", use_container_width=True):
                                _clean = _new_name.strip().upper()
                                if _clean and rename_company(_company, _clean):
                                    if _co_now == _company:
                                        st.session_state['_current_company'] = _clean
                                    st.session_state['_renaming_company'] = None
                                    st.success("✅ تم تغيير الاسم")
                                    _full_rerun()
                                else:
                                    st.error("❌ الاسم مستخدم بالفعل أو غير صالح")
                        with _rn_b:
                            if st.button("❌ إلغاء", key=f"_rename_cancel_{_company}", use_container_width=True):
                                st.session_state['_renaming_company'] = None
                                st.rerun()
                    elif st.session_state.get('_confirm_delete_company') == _company:
                        st.warning(f"⚠️ حذف **{_company}** بكل ملفاتها نهائيًا؟")
                        _dc1, _dc2 = st.columns(2)
                        with _dc1:
                            if st.button("✅ تأكيد الحذف", type="primary", key=f"_delco_yes_{_company}", use_container_width=True):
                                delete_company(_company)
                                if _co_now == _company:
                                    for _k in ('_current_company', '_current_year', '_current_month', '_cached_file_bytes'):
                                        st.session_state.pop(_k, None)
                                st.session_state['_confirm_delete_company'] = None
                                st.success("✅ تم حذف الشركة")
                                _full_rerun()
                        with _dc2:
                            if st.button("❌ إلغاء", key=f"_delco_no_{_company}", use_container_width=True):
                                st.session_state['_confirm_delete_company'] = None
                                st.rerun()
                    else:
                        if st.button("✏️ إعادة تسمية", key=f"_rename_btn_{_company}", use_container_width=True):
                            st.session_state['_renaming_company'] = _company
                            st.rerun()
                        if st.button("🗑️ حذف الشركة", key=f"_delco_btn_{_company}", use_container_width=True):
                            st.session_state['_confirm_delete_company'] = _company
                            st.rerun()

            if not st.session_state[_co_open_key]:
                continue

            # ── مستوى السنة (بدون Columns متداخلة — مسافة بادئة نصية بدل كده) ──
            for _year in sorted(_tree[_company].keys(), reverse=True):
                _yr_open_key = f"_tree_yr_open_{_company}_{_year}"
                if _yr_open_key not in st.session_state:
                    st.session_state[_yr_open_key] = (_company == _co_now and _year == _yr_now)

                _yr_arrow = ""
                if st.button(
                    f"{_IND1}{_yr_arrow} 📅 {_year}",
                    key=f"_yr_toggle_{_company}_{_year}", use_container_width=True
                ):
                    st.session_state[_yr_open_key] = not st.session_state[_yr_open_key]
                    st.rerun()

                if not st.session_state[_yr_open_key]:
                    continue

                # ── مستوى الشهر (ورقة/ملف) ──
                for _month in _tree[_company][_year]:
                    _mo_name = _ARABIC_MONTHS[_month]
                    _is_current = (_co_now == _company and _yr_now == _year and _mo_now == _month)
                    _mark = "🔵" if _is_current else "📄"

                    _mo_lbl, _mo_menu = st.columns([6, 1], gap="small")
                    with _mo_lbl:
                        if st.button(
                            f"{_IND2} {_mark} {_mo_name}" + (" (مفتوح)" if _is_current else ""),
                            key=f"_open_{_company}_{_year}_{_month}",
                            use_container_width=True, disabled=_is_current
                        ):
                            _open_month_file(_company, _year, _month)
                    with _mo_menu:
                        with st.popover("⋮", use_container_width=True):
                            if st.session_state.get('_confirm_delete_file') == (_company, _year, _month):
                                st.warning(f"⚠️ حذف ملف {_mo_name} {_year} نهائيًا؟")
                                _df1, _df2 = st.columns(2)
                                with _df1:
                                    if st.button("✅ تأكيد", type="primary", key=f"_delf_yes_{_company}_{_year}_{_month}", use_container_width=True):
                                        delete_month_file(_company, _year, _month)
                                        if _is_current:
                                            for _k in ('_current_company', '_current_year', '_current_month', '_cached_file_bytes'):
                                                st.session_state.pop(_k, None)
                                        st.session_state['_confirm_delete_file'] = None
                                        st.success("✅ تم حذف الملف")
                                        _full_rerun()
                                with _df2:
                                    if st.button("❌ إلغاء", key=f"_delf_no_{_company}_{_year}_{_month}", use_container_width=True):
                                        st.session_state['_confirm_delete_file'] = None
                                        st.rerun()
                            else:
                                if st.button("🗑️ حذف الملف", key=f"_delf_btn_{_company}_{_year}_{_month}", use_container_width=True):
                                    st.session_state['_confirm_delete_file'] = (_company, _year, _month)
                                    st.rerun()

            st.markdown(
                "<hr style='margin:0.2rem 0; opacity:0.1;'>", unsafe_allow_html=True
            )

    _render_file_tree()



    st.markdown("---")
    st.markdown("### 📤 فتح مؤقت (Anonymous)")
    st.caption("يفتح الملف للمراجعة بدون أي حفظ في بيانات البرنامج.")
    _anon_file = st.file_uploader(
        "اختر ملف XLS/XLSX", type=["xls", "xlsx"], key="_anon_uploader"
    )
    if _anon_file is not None and st.button("📤 فتح للمراجعة", key="_anon_open_btn", type="primary"):
        st.session_state['_cached_file_bytes'] = _anon_file.read()
        st.session_state['_current_company']  = None
        st.session_state['_current_year']     = None
        st.session_state['_current_month']    = None
        st.session_state['_is_anonymous']     = True
        st.session_state['corrections']       = {}
        st.session_state['day_overrides']     = {}
        st.session_state['saved_rates']       = {}
        st.session_state['restore_done']      = False
        st.session_state['file_state_loaded'] = False
        st.session_state['last_file_hash']    = None
        st.session_state['_show_import_popup'] = False
        st.rerun()

_show_upload_section = st.session_state.get('view_emp') is None

if _show_upload_section:

    # ── محاولة استعادة آخر جلسة تلقائياً ──
    if (st.session_state['_current_company'] is None
            and not st.session_state.get('_cached_file_bytes')):
        _last = load_last_session()
        if _last:
            _sess_bytes = load_work_file_bytes(
                _last['company'], _last['year'], _last['month']
            )
            if _sess_bytes:
                import datetime as _dt
                _mo_name = [
                    '', 'يناير','فبراير','مارس','أبريل','مايو','يونيو',
                    'يوليو','أغسطس','سبتمبر','أكتوبر','نوفمبر','ديسمبر'
                ][_last['month']]
                st.info(
                    f"📂 آخر ملف: **{_last['company']}** — {_mo_name} {_last['year']}"
                )
                _r1, _r2 = st.columns(2)
                with _r1:
                    if st.button("✅ فتح آخر ملف", type="primary", use_container_width=True):
                        st.session_state['_cached_file_bytes']  = _sess_bytes
                        st.session_state['_current_company']    = _last['company']
                        st.session_state['_current_year']       = _last['year']
                        st.session_state['_current_month']      = _last['month']
                        # تحميل أسعار الساعات الافتراضية
                        _defs = load_employee_defaults(_last['company'])
                        if _defs and not st.session_state.get('saved_rates'):
                            st.session_state['saved_rates'] = _defs
                        st.rerun()
                with _r2:
                    if st.button("📤 استيراد ملف جديد", use_container_width=True):
                        st.session_state['_show_import_popup'] = True
                        st.rerun()
                st.markdown("---")

    # ── Popup الاستيراد ──
    _companies = list_companies()

    if st.session_state.get('_show_import_popup') or not _companies:
        with st.container(border=True):
            st.markdown("### 📤 استيراد ملف حضور جديد")

            uploaded = st.file_uploader("اختر ملف XLS/XLSX", type=["xls", "xlsx"], key="new_import_uploader")

            _anon_import = st.checkbox(
                "📤 استيراد مؤقت (Anonymous) — بدون حفظ في بيانات البرنامج",
                value=False, key="_import_anon_checkbox"
            )

            if not _anon_import:
                _imp_col1, _imp_col2, _imp_col3 = st.columns(3)
                with _imp_col1:
                    _co_options = _companies + ["➕ شركة جديدة"] if _companies else ["➕ شركة جديدة"]
                    _co_sel = st.selectbox("🏢 الشركة", _co_options, key="_import_company_sel")
                    if _co_sel == "➕ شركة جديدة":
                        _company_input = st.text_input("اسم الشركة الجديدة", key="_new_company_name")
                    else:
                        _company_input = _co_sel
                with _imp_col2:
                    import datetime as _dt
                    _year_input = st.number_input(
                        "📅 السنة", min_value=2020, max_value=2100,
                        value=_dt.date.today().year, step=1, key="_import_year"
                    )
                with _imp_col3:
                    _month_input = st.number_input(
                        "📆 الشهر", min_value=1, max_value=12,
                        value=_dt.date.today().month, step=1, key="_import_month"
                    )
            else:
                st.caption("لن يُحفظ هذا الملف في بيانات البرنامج، وسيُغلق البرنامج تلقائيًا بعد إنهاء المراجعة.")
                _company_input = "ANON"  # قيمة placeholder، غير مستخدمة فعليًا في وضع Anonymous

            _btn_label = "📤 فتح للمراجعة" if _anon_import else "💾 حفظ واستيراد"
            _btn_imp, _btn_cancel = st.columns(2)
            with _btn_imp:
                _do_import = st.button(
                    _btn_label, type="primary", use_container_width=True,
                    disabled=(uploaded is None or (not _anon_import and not _company_input.strip()))
                )
            with _btn_cancel:
                if _companies and st.button("إلغاء", use_container_width=True):
                    st.session_state['_show_import_popup'] = False
                    st.rerun()

            if _do_import and uploaded and _anon_import:
                st.session_state['_cached_file_bytes'] = uploaded.read()
                st.session_state['_current_company']   = None
                st.session_state['_current_year']      = None
                st.session_state['_current_month']     = None
                st.session_state['_is_anonymous']      = True
                st.session_state['_show_import_popup'] = False
                st.session_state['restore_done']       = False
                st.session_state['file_state_loaded']  = False
                st.session_state['last_file_hash']     = None
                st.session_state['corrections']        = {}
                st.session_state['day_overrides']      = {}
                st.session_state['saved_rates']        = {}
                st.rerun()

            elif _do_import and uploaded and not _anon_import and _company_input.strip():
                _raw_bytes = uploaded.read()
                _co = _company_input.strip().upper()
                save_imported_file(_raw_bytes, _co, int(_year_input), int(_month_input))
                st.session_state['_cached_file_bytes'] = _raw_bytes
                st.session_state['_current_company']   = _co
                st.session_state['_current_year']      = int(_year_input)
                st.session_state['_current_month']     = int(_month_input)
                st.session_state['_is_anonymous']      = False
                st.session_state['_show_import_popup'] = False
                st.session_state['restore_done']       = False
                st.session_state['file_state_loaded']  = False
                st.session_state['last_file_hash']     = None
                st.session_state['corrections']        = {}
                st.session_state['day_overrides']      = {}
                # تحميل أسعار الساعات الافتراضية للشركة
                _defs = load_employee_defaults(_co)
                st.session_state['saved_rates'] = _defs if _defs else {}
                st.success(f"✅ تم حفظ الملف: {_co} — {int(_month_input):02d}/{int(_year_input)}")
                st.rerun()

        # لو مفيش ملف محفوظ ولا popup مفتوح نوقف
        if not st.session_state.get('_cached_file_bytes'):
            st.stop()

    elif not st.session_state.get('_cached_file_bytes'):
        # لا يوجد جلسة ولا popup — اعرض زر الاستيراد مباشرة
        st.info("👆 ابدأ باستيراد ملف الحضور")
        if st.button("📤 استيراد ملف جديد", type="primary"):
            st.session_state['_show_import_popup'] = True
            st.rerun()
        st.stop()

    # ── زر العودة للأصل في الشاشة الرئيسية ──
    _co  = st.session_state.get('_current_company')
    _yr  = st.session_state.get('_current_year')
    _mo  = st.session_state.get('_current_month')
    if _co and _yr and _mo:
        _top1, _top2, _top3 = st.columns([3, 1, 1])
        with _top2:
            if st.button("📤 استيراد ملف جديد", use_container_width=True):
                st.session_state['_show_import_popup'] = True
                st.rerun()
        with _top3:
            if st.button("🔄 العودة للأصل", use_container_width=True, type="secondary"):
                st.session_state['_confirm_reset'] = True
                st.rerun()

        # ── نافذة تأكيد العودة للأصل ──
        if st.session_state.get('_confirm_reset'):
            with st.container(border=True):
                st.warning(
                    "⚠️ **تحذير:** سيتم حذف جميع التعديلات والتصحيحات وأسعار الساعات "
                    f"للملف الحالي ({_co} — {_mo:02d}/{_yr}) والعودة للملف الأصلي. "
                    "**هذا الإجراء لا يمكن التراجع عنه.**"
                )
                _c1, _c2 = st.columns(2)
                with _c1:
                    if st.button("✅ نعم، ارجع للأصل", type="primary", use_container_width=True):
                        _orig = load_original_file_bytes(_co, _yr, _mo)
                        if _orig:
                            st.session_state['_cached_file_bytes'] = _orig
                            st.session_state['corrections']        = {}
                            st.session_state['day_overrides']      = {}
                            st.session_state['saved_rates']        = {}
                            st.session_state['fingerprint_types']  = {}
                            st.session_state['missing_hours']      = {}
                            st.session_state['restore_done']       = False
                            st.session_state['file_state_loaded']  = True
                            st.session_state['last_file_hash']     = None
                            st.session_state['_confirm_reset']     = False
                            save_work_file(_orig, _co, _yr, _mo)
                            save_file_state(_co, _yr, _mo, {
                                'corrections': {}, 'day_overrides': {},
                                'fingerprint_types': {}, 'hourly_rates': {},
                                'missing_hours': {},
                            })
                            st.success("✅ تم استعادة الملف الأصلي بنجاح")
                            st.rerun()
                        else:
                            st.error("❌ لم يتم العثور على النسخة الأصلية")
                with _c2:
                    if st.button("❌ إلغاء", use_container_width=True):
                        st.session_state['_confirm_reset'] = False
                        st.rerun()

    restore_file = None  # لم نعد نحتاجها (الاستعادة أصبحت تلقائية)

else:
    uploaded     = st.session_state.get('_cached_uploaded_file')
    restore_file = None

if not st.session_state.get('_cached_file_bytes'):
    st.info("👆 ارفع ملف XLS للبدء")
    st.stop()

file_bytes = st.session_state.get('_cached_file_bytes')
if not file_bytes:
    st.error("⚠️ لم يتم رفع ملف الحضور بعد. ارجع للصفحة الرئيسية وارفع الملف أولاً.")
    st.stop()

file_hash  = get_file_hash(file_bytes)

with st.spinner("جاري تحليل الملف..."):
    _device = st.session_state.get('_device_type_idx', 0)
    if _device == 1:  # HikVision
        try:
            df, emp_days = parse_file_hikvision(
                file_bytes, cutoff_hour=cutoff_hour, saturate_min=saturate_raw,
                duplicate_punch_tolerance=duplicate_punch_tolerance
            )
        except Exception as _hik_err:
            st.error(
                f"❌ تعذّر قراءة الملف كـ HikVision: {_hik_err}\n\n"
                "تأكد أن الملف يحتوي على ورقة باسم 'AttendanceRecord'، "
                "أو غيّر نوع الجهاز إلى Main Fingerprint."
            )
            st.stop()
    else:  # Main Fingerprint
        try:
            df, emp_days = parse_file(
                file_bytes, cutoff_hour=cutoff_hour, saturate_min=saturate_raw,
                duplicate_punch_tolerance=duplicate_punch_tolerance
            )
        except Exception as _main_err:
            st.error(
                f"❌ تعذّر قراءة الملف كـ Main Fingerprint: {_main_err}\n\n"
                "تأكد أن الملف يحتوي على ورقة باسم 'Att.log report'، "
                "أو غيّر نوع الجهاز إلى HikVision."
            )
            st.stop()

if 'corrections' not in st.session_state:
    st.session_state.corrections = {}
if 'day_overrides' not in st.session_state:
    st.session_state.day_overrides = {}
if 'saved_rates' not in st.session_state:
    st.session_state.saved_rates = {}
if 'fingerprint_types' not in st.session_state:
    st.session_state.fingerprint_types = {}
if 'missing_hours' not in st.session_state:
    st.session_state.missing_hours = {}
if 'last_file_hash' not in st.session_state:
    st.session_state.last_file_hash = None
if 'restore_done' not in st.session_state:
    st.session_state.restore_done = False
if 'file_state_loaded' not in st.session_state:
    st.session_state.file_state_loaded = False
if 'pending_changes' not in st.session_state:
    st.session_state.pending_changes = {}

# ── تحميل الحالة المحفوظة من ملف الشهر (إذا كان موجوداً) ──
if not st.session_state.file_state_loaded:
    current_company = st.session_state.get('_current_company')
    current_year = st.session_state.get('_current_year')
    current_month = st.session_state.get('_current_month')
    
    if current_company and current_year and current_month:
        saved_state = load_file_state(current_company, current_year, current_month)
        if saved_state:
            st.session_state.corrections = saved_state.get('corrections', {})
            st.session_state.day_overrides = saved_state.get('day_overrides', {})
            st.session_state.fingerprint_types = saved_state.get('fingerprint_types', {})
            st.session_state.saved_rates = saved_state.get('hourly_rates', {})
            st.session_state.missing_hours = saved_state.get('missing_hours', {})
            st.session_state.file_state_loaded = True

if restore_file is not None and not st.session_state.restore_done:
    restore_bytes = restore_file.read()
    restored_corrections, restored_rates, restored_day_overrides = load_corrections_from_excel(restore_bytes)
    if restored_corrections or restored_rates or restored_day_overrides:
        st.session_state.corrections   = restored_corrections
        st.session_state.saved_rates   = restored_rates
        st.session_state.day_overrides = restored_day_overrides
        st.session_state.restore_done  = True
        total_c = sum(len(v) for v in restored_corrections.values())
        st.success(
            f"✅ تم استرجاع {total_c} تصحيح، {len(restored_day_overrides)} تعديل يوم كامل، "
            f"و {len(restored_rates)} سعر ساعة بنجاح"
        )
    else:
        st.warning("⚠️ لم يتم العثور على بيانات محفوظة في هذا الملف")

if st.session_state.last_file_hash not in (None, file_hash):
    st.warning("⚠️ تم رفع ملف حضور مختلف. التصحيحات السابقة قد لا تنطبق على هذا الملف.")
st.session_state.last_file_hash = file_hash

overrides_summary = apply_overrides(emp_days)

df_display = df.copy()
for eid, ov in overrides_summary.items():
    mask = df_display['id'] == eid
    df_display.loc[mask, 'work_hours']      = ov['work_hours']
    df_display.loc[mask, 'work_minutes']    = ov['work_minutes']
    df_display.loc[mask, 'attendance_days'] = ov['attendance_days']
    df_display.loc[mask, 'absent_days']     = ov['absent_days']
    df_display.loc[mask, 'incomplete_days'] = ov['incomplete_days']

attended_df = df_display[df_display['work_hours'] > 0].copy()
all_df      = df_display.copy()

# نحسب الإجمالي بناءً على الموظفين "الحاضرين" فقط (اللي ظاهرين في القائمة)
total_incomplete = int(attended_df['incomplete_days'].sum())

# ── DEBUG: نطبع في الكونسول تفاصيل كل الأيام المحسوبة ضمن الـ16 ──────────
# ── DEBUG: نطبع في الكونسول تفاصيل كل الأيام المحسوبة ضمن الإجمالي ──────────
print("=" * 70)
print(f"[DEBUG] total_incomplete (حاضرين فقط) = {total_incomplete}")
print("[DEBUG] تفصيل incomplete_days لكل موظف حاضر:")
_debug_rows = attended_df[attended_df['incomplete_days'] > 0][['id', 'name', 'work_hours', 'attendance_days', 'incomplete_days']]
for _, _r in _debug_rows.iterrows():
    print(f"  - ID {_r['id']:>4} | {_r['name']:<15} | incomplete_days={int(_r['incomplete_days'])} | work_hours={_r['work_hours']}")
print(f"[DEBUG] مجموع incomplete_days للحاضرين = {_debug_rows['incomplete_days'].sum()}")
_absent_incomplete = df_display[(df_display['work_hours'] == 0) & (df_display['incomplete_days'] > 0)]['incomplete_days'].sum()
print(f"[DEBUG] (للمقارنة فقط) incomplete_days عند الغائبين تمامًا ومُستبعدة الآن = {_absent_incomplete}")
print("=" * 70)

# نعُدّ التصحيحات الفعلية فقط (نستبعد مفاتيح "_role" المرافقة لكل تصحيح)
resolved_keys = sum(
    1 for v in st.session_state.corrections.values()
    for k in v.keys() if not k.endswith('_role')
)
n_pending_global = count_pending_changes()

if 'view_emp' not in st.session_state:
    st.session_state.view_emp = None

if st.session_state.view_emp is None:
    col_s1, col_s2, col_s3 = st.columns(3)
    with col_s1:
        st.success(f"✅ {len(df)} موظف — {len(attended_df)} حضروا هذا الشهر")
    with col_s2:
        if total_incomplete > 0:
            st.warning(f"⚠️ {total_incomplete} يوم ببصمة ناقصة لم تُراجَع")
        else:
            st.success("✅ لا توجد بصمات ناقصة متبقية")
    with col_s3:
        if n_pending_global > 0:
            st.warning(f"⏳ {n_pending_global} تعديل معلّق بانتظار Apply")
        elif resolved_keys > 0:
            st.info(f"✏️ {resolved_keys} تصحيح مُطبَّق")
        else:
            st.info("ℹ️ لا توجد تعديلات معلّقة")

    if n_pending_global > 0:
        st.markdown(
            f"<div style='background:#fff3cd;border:1px solid #ffc107;"
            f"border-radius:8px;padding:10px 14px;margin:8px 0;'>"
            f"⏳ <b>عندك {n_pending_global} تعديل معلّق لم يُطبَّق بعد</b> — "
            f"لازم تعمل Apply قبل ما تعتبر البيانات نهائية.</div>",
            unsafe_allow_html=True,
        )
        if st.button("✅ Apply All Pending (تطبيق كل التعديلات المعلّقة)", type="primary", key="main_apply_all"):
            count = apply_all_pending()
            st.success(f"✅ تم تطبيق {count} تعديل بنجاح!")
            st.rerun()

min_sample_bulk = 3  # default, يُعاد تعريفه في الصفحة الرئيسية إذا ظهر الـ widget

# ── زر التطبيق الجماعي الذكي ─────────────────────────────────────────────────
if st.session_state.view_emp is None and total_incomplete > 0:
    st.markdown("---")
    bulk_col1, bulk_col2, bulk_col3 = st.columns([3, 2, 3])

    with bulk_col2:
        min_sample_bulk = st.number_input(
            "الحد الأدنى لعدد الأيام المشابهة",
            min_value=1, max_value=10, value=3, step=1
        )

    with bulk_col1:
        st.markdown("<div style='padding-top:28px'>", unsafe_allow_html=True)
        if st.button(
            "🤖 تطبيق الاقتراح الذكي لكل الموظفين",
            key="bulk_smart_all_btn",
            type="primary",
            use_container_width=True
        ):
            st.session_state['bulk_smart_preview'] = True
        st.markdown("</div>", unsafe_allow_html=True)

    with bulk_col3:
        st.markdown("<div style='padding-top:28px;'></div>", unsafe_allow_html=True)

    # ── تقرير المعاينة قبل التطبيق ───────────────────────────────────────
    if st.session_state.get('bulk_smart_preview'):
        result = bulk_smart_apply_all(emp_days, min_sample=int(min_sample_bulk))
        applied = result['applied']
        skipped = result['skipped']

        # إذا لم يُطبَّق شيء بعد (preview فقط) نعرض التقرير
        # لكن bulk_smart_apply_all يضيف للـ pending مباشرة عند الاستدعاء
        # لذا نعيد تصفير الـ pending ونعيد الحساب كـ dry-run أولاً
        # ← نتراجع عن الـ pending التي أضفناها للتو ونعرض التقرير أولاً
        # الحل: نستخدم flag منفصل — preview يحسب بدون إضافة، confirm يضيف

        # تراجع عن الإضافات التلقائية من bulk_smart_apply_all
        pend = st.session_state.get('pending_changes', {})
        for item in applied:
            eid_i  = item['eid']
            day_i  = item['day']
            # نحذف ما أضافه للتو
            keys_to_rm = [
                k for k in list(pend.keys())
                if k.startswith(f"corr_{eid_i}_")
                and pend[k].get('value') == item['suggest']
            ]
            for k in keys_to_rm:
                del pend[k]
        st.session_state.pending_changes = pend

        # عرض التقرير
        emp_names = {str(row['id']): row['name'] for _, row in df.iterrows()}

        st.markdown("### 📋 تقرير المعاينة")
        ta, ts = st.columns(2)
        with ta:
            st.success(f"✅ **{len(applied)}** بصمة ستُطبَّق عليها الاقتراح الذكي")
        with ts:
            st.warning(f"⏭️ **{len(skipped)}** بصمة ستُتجاهَل (لا تستوفي الشروط)")

        if applied:
            with st.expander(f"✅ البصمات التي ستُطبَّق ({len(applied)})", expanded=True):
                rows_a = []
                for item in applied:
                    punch_label = "حضور → انصراف مقترح" if item['type'] == 'check_in' else "انصراف → حضور مقترح"
                    rows_a.append({
                        'الموظف':      emp_names.get(item['eid'], item['eid']),
                        'اليوم':       item['day'],
                        'البصمة':      to_12h(item['ci']),
                        'النوع':       punch_label,
                        'الاقتراح':    to_12h(item['suggest']),
                        'عينة':        f"{item['sample']} يوم",
                    })
                st.dataframe(pd.DataFrame(rows_a), use_container_width=True, hide_index=True)

        if skipped:
            with st.expander(f"⏭️ البصمات المتجاهَلة ({len(skipped)})", expanded=False):
                rows_s = []
                for item in skipped:
                    rows_s.append({
                        'الموظف':  emp_names.get(item['eid'], item['eid']),
                        'اليوم':   item['day'],
                        'البصمة':  to_12h(item['ci']),
                        'السبب':   item['reason'],
                    })
                st.dataframe(pd.DataFrame(rows_s), use_container_width=True, hide_index=True)

        if applied:
            conf_col1, conf_col2, _ = st.columns([2, 2, 4])
            with conf_col1:
                if st.button(
                    f"✅ تأكيد وتطبيق {len(applied)} تصحيح",
                    key="bulk_smart_confirm",
                    type="primary",
                    use_container_width=True
                ):
                    # الآن نطبّق فعلاً
                    final = bulk_smart_apply_all(emp_days, min_sample=int(min_sample_bulk))
                    apply_all_pending()
                    st.session_state['bulk_smart_preview'] = False
                    st.success(f"✅ تم تطبيق {len(final['applied'])} تصحيح ذكي بنجاح!")
                    st.rerun()
            with conf_col2:
                if st.button("❌ إلغاء", key="bulk_smart_cancel", use_container_width=True):
                    st.session_state['bulk_smart_preview'] = False
                    st.rerun()
        else:
            st.info("لا توجد بصمات تستوفي شروط التطبيق الآمن.")
            if st.button("إغلاق", key="bulk_smart_close"):
                st.session_state['bulk_smart_preview'] = False
                st.rerun()

    st.markdown("---")

if 'view_emp' not in st.session_state:
    st.session_state.view_emp = None

if st.session_state.view_emp is not None:
    eid     = st.session_state.view_emp
    emp_row = df_display[df_display['id'] == eid]
    if not emp_row.empty:
        col_back, col_info = st.columns([1, 5])
        with col_back:
            if st.button("← رجوع للقائمة الرئيسية"):
                # تحقق من تعديلات معلّقة قبل المغادرة
                n = count_pending_changes()
                if n > 0:
                    st.session_state['leaving_to_main'] = True
                else:
                    st.session_state.view_emp = None
                    st.rerun()

        # Dialog تأكيد المغادرة
        if st.session_state.get('leaving_to_main'):
            n = count_pending_changes()
            st.warning(f"⚠️ لديك {n} تعديل لم يُطبَّق بعد!")
            dc1, dc2, dc3 = st.columns(3)
            with dc1:
                if st.button("💾 حفظ وانتقال", type="primary", key="leave_save"):
                    apply_all_pending()
                    st.session_state.view_emp = st.session_state.pop('pending_target_emp', None)
                    st.session_state['leaving_to_main'] = False
                    st.rerun()
            with dc2:
                if st.button("❌ إلغاء (ابقَ في الصفحة)", key="leave_cancel"):
                    st.session_state['leaving_to_main'] = False
                    st.session_state.pop('pending_target_emp', None)
                    st.rerun()
            with dc3:
                if st.button("🗑️ تجاهل وانتقال", key="leave_discard"):
                    discard_all_pending()
                    st.session_state.view_emp = st.session_state.pop('pending_target_emp', None)
                    st.session_state['leaving_to_main'] = False
                    st.rerun()
        else:
            show_employee_detail(emp_row.iloc[0], emp_days, overrides_summary)
        st.stop()

st.markdown("## ⚙️ إعدادات الرواتب")

show_all = st.checkbox("عرض الموظفين الغائبين كل الشهر أيضاً", value=False)
edit_df  = all_df if show_all else attended_df

st.markdown("### 💰 سعر الساعة لكل موظف")
hourly_rates = {}
rate_cols    = st.columns(3)

for i, (_, emp) in enumerate(edit_df.iterrows()):
    eid        = str(emp['id'])
    saved_rate = st.session_state.saved_rates.get(eid, 0.0)
    with rate_cols[i % 3]:
        c_name, c_btn = st.columns([3, 1])
        with c_name:
            warn = f"  |  ⚠️ {int(emp['incomplete_days'])} ناقصة" if emp['incomplete_days'] > 0 else ""
            st.markdown(
                f"**{emp['name']}** (ID: {eid})  \n"
                f"🕐 {fmt_h(emp['work_hours'])}  |  📅 {fmt_days(emp['attendance_days'])} يوم{warn}"
            )
        with c_btn:
            if st.button("🔍", key=f"view_{eid}"):
                if count_pending_changes() > 0:
                    st.session_state['pending_target_emp'] = eid
                    st.session_state['leaving_to_main'] = True
                else:
                    st.session_state.view_emp = eid
                    st.session_state['leaving_to_main'] = False
                st.rerun()
        rate = st.number_input(
            "سعر الساعة", min_value=0.0, value=saved_rate, step=1.0,
            key=f"rate_{eid}", label_visibility="collapsed"
        )
        hourly_rates[eid] = rate
        # حفظ السعر في session_state فوراً
        if rate > 0:
            st.session_state.saved_rates[eid] = rate

# حفظ أسعار الساعات تلقائياً كإعدادات افتراضية للشركة + حفظ الحالة الكاملة
_curr_co = st.session_state.get('_current_company')
if _curr_co and hourly_rates:
    save_employee_defaults(_curr_co, hourly_rates)
    # ✅ حفظ الحالة الكاملة (بما فيها الأسعار والتصحيحات والتعديلات)
    auto_save_file_state()

payroll_df = calculate_payroll(edit_df, hourly_rates, overrides_summary)

st.markdown("---")
st.markdown("## 📥 تحميل الملفات")
dl_col1, dl_col2 = st.columns(2)

with dl_col1:
    st.markdown("### 📊 كشف الرواتب فقط")
    if not payroll_df.empty:
        excel_buf = export_to_excel(payroll_df)
        st.download_button(
            label="⬇️ تحميل كشف الرواتب",
            data=excel_buf,
            file_name="payroll_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.button("⬇️ تحميل كشف الرواتب", disabled=True)

with dl_col2:
    st.markdown("### 📦 الملف النهائي الكامل")
    full_buf = export_full_file(
        payroll_df if not payroll_df.empty else pd.DataFrame(),
        st.session_state.corrections,
        hourly_rates,
        st.session_state.day_overrides
    )
    st.download_button(
        label="⬇️ تحميل الملف النهائي الكامل",
        data=full_buf,
        file_name="final_complete.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary"
    )

if payroll_df.empty:
    st.warning("⚠️ أدخل سعر الساعة لموظف واحد على الأقل لعرض كشف الرواتب")
    st.stop()

st.markdown("---")
st.markdown("## 📊 إحصائيات عامة")
m1, m2, m3, m4 = st.columns(4)
with m1: st.metric("👥 عدد الموظفين",        len(payroll_df))
with m2: st.metric("💵 إجمالي الرواتب",      f"{payroll_df['صافي الراتب'].sum():,.2f}")
with m3: st.metric("⏱️ إجمالي ساعات العمل",  f"{payroll_df['ساعات العمل الفعلية'].sum():,.1f}")
with m4: st.metric("⚠️ ببصمات ناقصة",       int((payroll_df['أيام بصمة ناقصة'] > 0).sum()))

st.markdown("---")
st.markdown("## 🔍 البحث والتصفية")
s_col, mn_col, mx_col = st.columns([2, 1, 1])
with s_col:
    search = st.text_input("🔎 ابحث عن موظف بالاسم أو الـ ID")
with mn_col:
    min_sal = st.number_input("الحد الأدنى للراتب", value=0.0, step=100.0)
with mx_col:
    max_sal = st.number_input("الحد الأقصى للراتب",
                               value=float(payroll_df['صافي الراتب'].max()) + 1, step=100.0)

filtered = payroll_df.copy()
if search:
    mask = (filtered['الاسم'].str.contains(search, case=False, na=False) |
            filtered['ID'].str.contains(search, case=False, na=False))
    filtered = filtered[mask]
filtered = filtered[(filtered['صافي الراتب'] >= min_sal) & (filtered['صافي الراتب'] <= max_sal)]

st.markdown(f"## 📋 كشف الرواتب ({len(filtered)} موظف)")
st.dataframe(filtered, use_container_width=True, hide_index=True)

st.markdown("---")
st.markdown("## 📈 الرسوم البيانية")
tab1, tab2, tab3 = st.tabs(["💰 توزيع الرواتب", "⏱️ ساعات العمل", "📅 أيام الحضور"])
with tab1:
    if not filtered.empty:
        st.bar_chart(filtered.set_index('الاسم')['صافي الراتب'].sort_values(ascending=False))
with tab2:
    if not filtered.empty:
        st.bar_chart(filtered.set_index('الاسم')['ساعات العمل الفعلية'])
with tab3:
    if not filtered.empty:
        st.bar_chart(filtered[['الاسم', 'أيام الحضور', 'أيام الغياب']].set_index('الاسم'))